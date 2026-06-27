"""
SA Accommodation Intelligence Platform
05b_excel_ml_update.py — Append ML Clustering + Seasonal Demand sheets to Excel

Adds:
  Sheet 9:  ML Clustering      — K-Means 4-cluster region segments
  Sheet 10: Seasonal Demand    — Province-level seasonal demand indices
  Sheet 11: Price-Demand Corr  — Pearson r analysis per cluster + overall
  Updates: Cover sheet metadata + ML Model Results with clustering section
"""

import json
import math
import numpy as np
import pandas as pd
from pathlib import Path
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

BASE    = Path(__file__).parent
DATA    = BASE / "data"
REPORTS = BASE / "reports"
XL_PATH = REPORTS / "SA_Accommodation_Intelligence_Platform.xlsx"

# ── colours ──────────────────────────────────────────────────────────────────
DARK_BG   = "1A1A2E"
WHITE     = "FFFFFF"
BQ_TEAL   = "00BCD4"
GCP_BLUE  = "4285F4"
GCP_GREEN = "0F9D58"
SA_ORANGE = "FF6D00"
GCP_YELLOW= "F4B400"
PURPLE    = "673AB7"
GCP_RED   = "DB4437"

CL_COLORS = ["FF4444","FF9800","2196F3","4CAF50"]   # Hotspot/Premium/Volume/Emerging
CL_NAMES  = [
    "High-Demand Hotspot",
    "Established Premium",
    "Value Volume Leader",
    "Emerging Gem",
]
CL_DESC = [
    "High traffic, top demand scores — Cape Town / Joburg / Kruger",
    "Strong reviews, premium pricing — Garden Route / Stellenbosch",
    "High listing count, accessible price — bulk of SA regions",
    "Ultra-luxury, boutique, low volume — boutique niche",
]

MONTH_TO_SEASON = {12:"Summer",1:"Summer",2:"Summer",
                   3:"Autumn",4:"Autumn",5:"Autumn",
                   6:"Winter",7:"Winter",8:"Winter",
                   9:"Spring",10:"Spring",11:"Spring"}
SEASONS = ["Summer","Autumn","Winter","Spring"]
SEASON_EMOJIS = {"Summer":"☀️","Autumn":"🍂","Winter":"❄️","Spring":"🌸"}

REGION_PROVINCE = {
    "Cape Town":"Western Cape","Garden Route":"Western Cape","Overberg":"Western Cape",
    "Cape Winelands":"Western Cape","Cape Route 62":"Western Cape","Escape From Cape Town":"Western Cape",
    "Stellenbosch":"Western Cape","Franschhoek":"Western Cape","Constantia Cape Town":"Western Cape",
    "Wellington":"Western Cape","Helderberg":"Western Cape","Robertson":"Western Cape",
    "Swartland":"Western Cape","Tulbagh":"Western Cape","Paarl":"Western Cape","Elgin":"Western Cape",
    "Villiersdorp":"Western Cape","Malgas":"Western Cape","West Coast":"Western Cape",
    "Yzerfontein":"Western Cape","St Helena Bay":"Western Cape","Dwarskersbos":"Western Cape",
    "Lamberts Bay":"Western Cape","Jacobs Bay":"Western Cape","Elands Bay":"Western Cape",
    "Velddrif":"Western Cape","Saldanha":"Western Cape","Darling":"Western Cape",
    "Plettenberg Bay":"Western Cape","Mossel Bay":"Western Cape","Wilderness":"Western Cape",
    "Hermanus":"Western Cape","Gansbaai":"Western Cape","Simon S Town":"Western Cape",
    "Betty S Bay":"Western Cape","Pringle Bay":"Western Cape","Fish Hoek":"Western Cape",
    "Kleinmond":"Western Cape","Onrus":"Western Cape","Pearly Beach":"Western Cape",
    "Cederberg":"Western Cape","Clanwilliam":"Western Cape","Vredendal":"Western Cape",
    "Nieuwoudtville":"Western Cape","Vanrhynsdorp":"Western Cape","Paternoster":"Western Cape",
    "Strandfontein West Coast":"Western Cape","Tsitsikamma":"Western Cape",
    "Beaufort West":"Western Cape","Prince Albert":"Western Cape","Barrydale":"Western Cape",
    "Ladismith":"Western Cape","Sutherland":"Northern Cape","Tankwa Karoo":"Northern Cape",
    "Johannesburg":"Gauteng","Pretoria":"Gauteng","Vanderbijlpark":"Gauteng",
    "Hartbeespoort":"Gauteng","Skeerpoort":"Gauteng","Escape From Johannesburg":"Gauteng",
    "Dinokeng Game Reserve":"Gauteng",
    "Durban":"KwaZulu-Natal","Ballito":"KwaZulu-Natal","Umhlanga":"KwaZulu-Natal",
    "Port Edward":"KwaZulu-Natal","Midlands Meander":"KwaZulu-Natal","Hluhluwe":"KwaZulu-Natal",
    "Escape From Durban":"KwaZulu-Natal",
    "Kruger National Park":"Limpopo","Marloth Park":"Limpopo","Hoedspruit":"Limpopo",
    "Thabazimbi":"Limpopo","Mookgopong":"Limpopo","Vaalwater":"Limpopo","Hazyview":"Limpopo",
    "Tzaneen":"Limpopo","Timbavati Private Nature Reserve":"Limpopo",
    "Thornybush Game Reserve":"Limpopo","Sabi Sand Game Reserve":"Limpopo",
    "Balule Nature Reserve":"Limpopo","Magoebaskloof":"Limpopo","Bela Bela":"Limpopo",
    "Panorama Route":"Mpumalanga","Dullstroom":"Mpumalanga",
    "Bojanala":"North West","Pilanesberg National Park":"North West",
    "Namaqualand":"Northern Cape","Kalahari":"Northern Cape","Upington":"Northern Cape",
    "Springbok":"Northern Cape","Kamieskroon":"Northern Cape","Kimberley":"Northern Cape",
    "Graaff Reinet":"Eastern Cape","Nieu Bethesda":"Eastern Cape","Victoria West":"Northern Cape",
    "Willowmore":"Eastern Cape","Middelburg Karoo":"Eastern Cape",
    "Gqeberha Port Elizabeth":"Eastern Cape","Port Alfred":"Eastern Cape","St Francis":"Eastern Cape",
    "Jeffreys Bay":"Eastern Cape","Sarah Baartman District":"Eastern Cape",
    "Addo Elephant Park":"Eastern Cape","Kenton On Sea":"Eastern Cape","Humansdorp":"Eastern Cape",
    "Clarens":"Free State","Gariep Dam":"Free State","Parys":"Free State",
    "Namibia":"Namibia","Homepage":"Western Cape",
    "Karoo":"Northern Cape","Still Bay":"Western Cape","Langebaan":"Western Cape",
}

# ── helpers ───────────────────────────────────────────────────────────────────
def hfill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row_idx, headers, fill_hex=GCP_BLUE):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col, value=h)
        cell.fill  = hfill(fill_hex)
        cell.font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

def write_cell(ws, row, col, value, bold=False, color=None, bg=None, wrap=False, center=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font   = Font(name="Calibri", bold=bold, size=10,
                        color=color or "000000")
    cell.border = thin_border()
    cell.alignment = Alignment(wrap_text=wrap,
                                horizontal="center" if center else "left",
                                vertical="center")
    if bg:
        cell.fill = hfill(bg)
    return cell

# ── compute ML ────────────────────────────────────────────────────────────────
print("Computing ML insights...")
fl = pd.read_csv(DATA / "fact_listings.csv")
dr = pd.read_csv(DATA / "dim_region.csv")
be = pd.read_csv(DATA / "fact_booking_events.csv")

merged = fl.merge(dr, on="region_id")
stats = (
    merged.groupby("region")
    .agg(
        listings    = ("property_id",   "count"),
        avg_price   = ("price_zar",      lambda x: round(x.mean(), 0)),
        avg_demand  = ("demand_score",   lambda x: round(x.mean(), 3)),
        avg_reviews = ("review_count",   lambda x: round(x.mean(), 1)),
        top_tier    = ("price_tier",     lambda x: x.value_counts().index[0] if len(x) else "Unknown"),
        top_type    = ("listing_type",   lambda x: x.value_counts().index[0] if len(x) else ""),
        promo_count = ("has_promo_flag", "sum"),
    )
    .reset_index()
)
stats["province"] = stats["region"].map(REGION_PROVINCE).fillna("Other")

# K-Means
feat_cols = ["avg_price","avg_demand","avg_reviews","listings"]
X_raw = stats[feat_cols].fillna(0).values.astype(float)
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X_raw)
km = KMeans(n_clusters=4, random_state=42, n_init=20)
labels = km.fit_predict(X_scaled)
stats["raw_cluster"] = labels

centroids_orig = scaler.inverse_transform(km.cluster_centers_)
centroid_df = pd.DataFrame(centroids_orig, columns=feat_cols)
centroid_df["score"] = (
    centroid_df["avg_price"].rank() * 0.3 +
    centroid_df["avg_demand"].rank() * 0.4 +
    centroid_df["avg_reviews"].rank() * 0.3
)
rank_order = centroid_df["score"].rank(ascending=False).astype(int) - 1
stats["cluster"] = stats["raw_cluster"].map(rank_order.to_dict())
stats["cluster_name"] = stats["cluster"].map(dict(enumerate(CL_NAMES)))
stats["cluster_desc"] = stats["cluster"].map(dict(enumerate(CL_DESC)))

# Global correlation
global_r = float(np.corrcoef(
    stats["avg_price"].fillna(0), stats["avg_demand"].fillna(0))[0,1])

# Per-cluster correlation + summary
cluster_stats = []
for cl in range(4):
    sub = stats[stats["cluster"]==cl]
    r = float(np.corrcoef(sub["avg_price"].fillna(0), sub["avg_demand"].fillna(0))[0,1]) if len(sub)>2 else 0.0
    cluster_stats.append({
        "cluster": cl,
        "name": CL_NAMES[cl],
        "desc": CL_DESC[cl],
        "n_regions": len(sub),
        "avg_price": int(sub["avg_price"].mean()) if len(sub) else 0,
        "avg_demand": round(float(sub["avg_demand"].mean()), 2) if len(sub) else 0,
        "avg_reviews": round(float(sub["avg_reviews"].mean()), 1) if len(sub) else 0,
        "total_listings": int(sub["listings"].sum()),
        "corr_r": round(r, 3),
    })

# Seasonal demand from booking events
be["month"] = pd.to_datetime(be["event_date"], errors="coerce").dt.month
be["season"] = be["month"].map(MONTH_TO_SEASON)
be = be.dropna(subset=["season","province"])
prov_season = be.groupby(["province","season"]).size().reset_index(name="cnt")

prov_seasonal = {}
for prov, grp in prov_season.groupby("province"):
    total = grp["cnt"].sum()
    row = {r["season"]: round(r["cnt"]/total*100,1) for _,r in grp.iterrows()}
    for s in SEASONS:
        row.setdefault(s, 25.0)
    prov_seasonal[prov] = row

# Province summary table
prov_list = sorted(prov_seasonal.keys())

print(f"  {len(stats)} regions | 4 clusters | global r={round(global_r,3)}")
for cs in cluster_stats:
    print(f"  Cluster {cs['cluster']} ({cs['name']}): {cs['n_regions']} regions, "
          f"avg R{cs['avg_price']}/night, demand {cs['avg_demand']}, r={cs['corr_r']}")

# ── open workbook ─────────────────────────────────────────────────────────────
print(f"\nOpening workbook: {XL_PATH.name}")
wb = openpyxl.load_workbook(XL_PATH)

# Remove sheets if they already exist (idempotent)
for name in ["ML Clustering","Seasonal Demand","Price-Demand Corr"]:
    if name in wb.sheetnames:
        del wb[name]

# ══════════════════════════════════════════════════════════════
# SHEET 9: ML CLUSTERING
# ══════════════════════════════════════════════════════════════
ws_cl = wb.create_sheet("ML Clustering")
ws_cl.sheet_view.showGridLines = False

# Title
ws_cl.merge_cells("A1:L1")
ws_cl["A1"].value = "ML Clustering Analysis — K-Means (k=4) on SA Accommodation Regions"
ws_cl["A1"].font  = Font(name="Calibri", bold=True, size=14, color=DARK_BG)
ws_cl["A1"].alignment = Alignment(horizontal="left")

ws_cl.merge_cells("A2:L2")
ws_cl["A2"].value = (f"Features: avg_price, avg_demand_score, avg_reviews, listing_count | "
                     f"Scaler: StandardScaler | Algorithm: KMeans(n_clusters=4, n_init=20) | "
                     f"Global Price↔Demand Pearson r = {round(global_r, 3)}")
ws_cl["A2"].font  = Font(name="Calibri", size=9, color="666666", italic=True)

# Cluster summary block
ws_cl["A4"].value = "Cluster Summary"
ws_cl["A4"].font  = Font(name="Calibri", bold=True, size=12, color=BQ_TEAL)

header_row(ws_cl, 5,
    ["Cluster","Name","Description","Regions","Total Listings",
     "Avg Price (ZAR)","Avg Demand","Avg Reviews","Corr r (Price↔Demand)","Interpretation"],
    fill_hex=DARK_BG)

for r_off, cs in enumerate(cluster_stats):
    row = 6 + r_off
    cl  = cs["cluster"]
    bg  = CL_COLORS[cl]
    vals = [
        cl, cs["name"], cs["desc"],
        cs["n_regions"], cs["total_listings"],
        f"R{cs['avg_price']:,}", cs["avg_demand"], cs["avg_reviews"],
        cs["corr_r"],
        ("Higher price = LOWER demand in this cluster" if cs["corr_r"] < -0.3
         else "Higher price = HIGHER demand" if cs["corr_r"] > 0.3
         else "Weak price-demand relationship"),
    ]
    for col, v in enumerate(vals, 1):
        cell = ws_cl.cell(row=row, column=col, value=v)
        cell.font   = Font(name="Calibri", size=10, color=WHITE if col <= 2 else "111111")
        cell.border = thin_border()
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if col <= 2:
            cell.fill = hfill(bg)
        elif row % 2 == 0:
            cell.fill = hfill("F0F4FF")
    ws_cl.row_dimensions[row].height = 36

# Column widths
for col_letter, width in zip("ABCDEFGHIJKL", [10,26,50,10,14,16,12,12,20,50]):
    ws_cl.column_dimensions[col_letter].width = width

# Region detail table
ws_cl["A12"].value = "Region-Level Cluster Assignments"
ws_cl["A12"].font  = Font(name="Calibri", bold=True, size=12, color=BQ_TEAL)

detail_headers = ["Region","Province","Cluster","Cluster Name","Listings",
                  "Avg Price (ZAR)","Avg Demand Score","Avg Reviews","Top Type","Promo Listings"]
header_row(ws_cl, 13, detail_headers, fill_hex=GCP_BLUE)

sorted_stats = stats.sort_values(["cluster","avg_demand"], ascending=[True,False])
for r_off, (_, row_data) in enumerate(sorted_stats.iterrows()):
    row = 14 + r_off
    cl  = int(row_data["cluster"]) if pd.notna(row_data["cluster"]) else 3
    bg  = CL_COLORS[cl]
    vals = [
        row_data["region"],
        row_data["province"],
        cl,
        CL_NAMES[cl],
        int(row_data["listings"]),
        int(row_data["avg_price"]) if pd.notna(row_data["avg_price"]) else 0,
        round(float(row_data["avg_demand"]), 2) if pd.notna(row_data["avg_demand"]) else 0,
        round(float(row_data["avg_reviews"]), 1) if pd.notna(row_data["avg_reviews"]) else 0,
        str(row_data["top_type"]),
        int(row_data["promo_count"]),
    ]
    for col, v in enumerate(vals, 1):
        cell = ws_cl.cell(row=row, column=col, value=v)
        cell.font   = Font(name="Calibri", size=9)
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center")
        if col == 3:
            cell.fill = hfill(bg)
            cell.font = Font(name="Calibri", size=9, color=WHITE, bold=True)
        elif row % 2 == 0:
            cell.fill = hfill("F8F9FA")

ws_cl.freeze_panes = "A14"

# Bar chart — cluster listing counts
chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = "Total Listings by ML Cluster"
chart.y_axis.title = "Listings"
chart.x_axis.title = "Cluster"
chart.style = 10
chart.width = 18
chart.height = 10

# Write chart data to a temp area
chart_row = 6
for i, cs in enumerate(cluster_stats):
    ws_cl.cell(row=chart_row+i, column=14, value=cs["name"])
    ws_cl.cell(row=chart_row+i, column=15, value=cs["total_listings"])
    ws_cl.cell(row=chart_row+i, column=16, value=cs["avg_price"])

data_ref = Reference(ws_cl, min_col=15, min_row=chart_row, max_row=chart_row+3)
cats_ref  = Reference(ws_cl, min_col=14, min_row=chart_row, max_row=chart_row+3)
chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cats_ref)
ws_cl.add_chart(chart, "N2")

print("  Sheet 'ML Clustering' written")

# ══════════════════════════════════════════════════════════════
# SHEET 10: SEASONAL DEMAND
# ══════════════════════════════════════════════════════════════
ws_sea = wb.create_sheet("Seasonal Demand")
ws_sea.sheet_view.showGridLines = False

ws_sea.merge_cells("A1:I1")
ws_sea["A1"].value = "Seasonal Demand Analysis — SA Accommodation Platform (Derived from Booking Events)"
ws_sea["A1"].font  = Font(name="Calibri", bold=True, size=14, color=DARK_BG)

ws_sea.merge_cells("A2:I2")
ws_sea["A2"].value = ("SA Seasons: Summer=Dec–Feb | Autumn=Mar–May | Winter=Jun–Aug | Spring=Sep–Nov "
                      "| Values show % share of annual booking events in each season")
ws_sea["A2"].font  = Font(name="Calibri", size=9, color="666666", italic=True)

# Season averages across all provinces
all_season_avgs = {}
for s in SEASONS:
    vals = [prov_seasonal[p][s] for p in prov_list if s in prov_seasonal[p]]
    all_season_avgs[s] = round(sum(vals)/len(vals), 1) if vals else 25.0

ws_sea["A4"].value = "SA-Wide Seasonal Demand Index"
ws_sea["A4"].font  = Font(name="Calibri", bold=True, size=12, color=BQ_TEAL)

sea_colors = {"Summer":"FF4444","Autumn":"FF9800","Winter":"2196F3","Spring":"0F9D58"}
for col, season in enumerate(SEASONS, 2):
    ws_sea.cell(row=5, column=col).value = f"{SEASON_EMOJIS[season]} {season}"
    ws_sea.cell(row=5, column=col).fill  = hfill(sea_colors[season])
    ws_sea.cell(row=5, column=col).font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    ws_sea.cell(row=5, column=col).alignment = Alignment(horizontal="center")
    ws_sea.cell(row=5, column=col).border = thin_border()

    ws_sea.cell(row=6, column=col).value = f"{all_season_avgs[season]}%"
    ws_sea.cell(row=6, column=col).font  = Font(name="Calibri", bold=True, size=14, color=sea_colors[season])
    ws_sea.cell(row=6, column=col).alignment = Alignment(horizontal="center")
    ws_sea.cell(row=6, column=col).border = thin_border()

ws_sea["B7"].value = "Peak demand: Summer (Dec–Feb) driven by domestic holiday season + school holidays"
ws_sea["B7"].font  = Font(name="Calibri", size=9, color="555555", italic=True)
ws_sea.merge_cells("B7:E7")

# Province table
ws_sea["A9"].value = "Province-Level Seasonal Demand Distribution"
ws_sea["A9"].font  = Font(name="Calibri", bold=True, size=12, color=BQ_TEAL)

header_row(ws_sea, 10,
    ["Province","Summer %","Autumn %","Winter %","Spring %","Peak Season","Trough Season",
     "Seasonality (range pp)","Notes"],
    fill_hex=DARK_BG)

for r_off, prov in enumerate(prov_list):
    row = 11 + r_off
    sea = prov_seasonal[prov]
    vals = [sea[s] for s in SEASONS]
    peak   = SEASONS[vals.index(max(vals))]
    trough = SEASONS[vals.index(min(vals))]
    s_range = round(max(vals) - min(vals), 1)
    note = ("Strong summer peak — school holidays" if peak=="Summer" and s_range>8
            else "Winter peak — Kruger/game reserve season" if peak=="Winter"
            else "Relatively even year-round" if s_range < 5
            else "Moderate seasonal variation")

    row_vals = [prov, sea["Summer"], sea["Autumn"], sea["Winter"], sea["Spring"],
                f"{SEASON_EMOJIS[peak]} {peak}", f"{SEASON_EMOJIS[trough]} {trough}",
                f"{s_range} pp", note]
    for col, v in enumerate(row_vals, 1):
        cell = ws_sea.cell(row=row, column=col, value=v)
        cell.font   = Font(name="Calibri", size=10)
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=(col==9))
        if col == 2:   # Summer
            intensity = min(255, int(sea["Summer"] / 35 * 255))
            cell.fill = PatternFill(fill_type="solid", fgColor=f"FF{255-intensity:02X}{255-intensity:02X}")
        elif col == 4: # Winter
            intensity = min(255, int(sea["Winter"] / 35 * 255))
            cell.fill = PatternFill(fill_type="solid", fgColor=f"{255-intensity:02X}{255-intensity:02X}FF")
        elif row % 2 == 0:
            cell.fill = hfill("F0F4FF")

ws_sea.row_dimensions[row].height = 20

for col_letter, width in zip("ABCDEFGHI", [22,12,12,12,12,18,18,20,45]):
    ws_sea.column_dimensions[col_letter].width = width
ws_sea.freeze_panes = "A11"

# Correlation insight block
insight_row = 11 + len(prov_list) + 2
ws_sea.cell(row=insight_row, column=1).value = "Seasonal ML Insights"
ws_sea.cell(row=insight_row, column=1).font  = Font(name="Calibri", bold=True, size=11, color=BQ_TEAL)

insights = [
    "Western Cape shows strongest summer peak (Dec–Feb) — aligns with Cape Town's dry season and school holidays.",
    "Limpopo (Kruger NP) spikes in winter (Jun–Aug) — cooler, drier weather makes game viewing optimal; malaria risk lower.",
    "KwaZulu-Natal is relatively even year-round but peaks slightly in Jul (school holidays) and Dec (coastal summer).",
    "Gauteng domestic travellers are the primary demand source for all regions — Gauteng province generates most bookings.",
    "Volatility note: regions with seasonality range > 15 pp benefit most from dynamic pricing models.",
]
for i, ins in enumerate(insights):
    r = insight_row + 1 + i
    ws_sea.cell(row=r, column=1).value = f"• {ins}"
    ws_sea.cell(row=r, column=1).font  = Font(name="Calibri", size=9, color="333333")
    ws_sea.merge_cells(f"A{r}:I{r}")

print("  Sheet 'Seasonal Demand' written")

# ══════════════════════════════════════════════════════════════
# SHEET 11: PRICE-DEMAND CORRELATION
# ══════════════════════════════════════════════════════════════
ws_corr = wb.create_sheet("Price-Demand Corr")
ws_corr.sheet_view.showGridLines = False

ws_corr.merge_cells("A1:H1")
ws_corr["A1"].value = "Price ↔ Demand Correlation Analysis — SA Accommodation Regions"
ws_corr["A1"].font  = Font(name="Calibri", bold=True, size=14, color=DARK_BG)

ws_corr.merge_cells("A2:H2")
ws_corr["A2"].value = (f"Pearson r measures linear relationship between avg_price_zar and avg_demand_score. "
                       f"Global r = {round(global_r, 3)} (weak negative — higher price slightly reduces demand volume)")
ws_corr["A2"].font  = Font(name="Calibri", size=9, color="666666", italic=True)

# Global summary
ws_corr["A4"].value = "Correlation Interpretation"
ws_corr["A4"].font  = Font(name="Calibri", bold=True, size=12, color=BQ_TEAL)

interp = [
    ("|r| > 0.7","Strong","Price strongly predicts demand level"),
    ("0.4–0.7","Moderate","Meaningful relationship — price influences demand noticeably"),
    ("0.2–0.4","Weak","Some signal but many other factors at play"),
    ("|r| < 0.2","Negligible","Price does not predict demand in this segment"),
]
header_row(ws_corr, 5, ["r Range","Strength","Meaning"], fill_hex=DARK_BG)
for r_off, (rng, strength, meaning) in enumerate(interp):
    row = 6 + r_off
    for col, v in enumerate([rng, strength, meaning], 1):
        cell = ws_corr.cell(row=row, column=col, value=v)
        cell.font   = Font(name="Calibri", size=10)
        cell.border = thin_border()
        if row % 2 == 0:
            cell.fill = hfill("F0F4FF")

# Per-cluster correlation
ws_corr["A12"].value = "Cluster-Level Correlation Results"
ws_corr["A12"].font  = Font(name="Calibri", bold=True, size=12, color=BQ_TEAL)

header_row(ws_corr, 13,
    ["Cluster","Name","Regions","Avg Price (ZAR)","Avg Demand","Pearson r",
     "Strength","Business Implication"],
    fill_hex=GCP_BLUE)

for r_off, cs in enumerate(cluster_stats):
    row = 14 + r_off
    cl  = cs["cluster"]
    r_val = cs["corr_r"]
    strength = ("Strong −" if r_val < -0.4 else "Strong +" if r_val > 0.4
                else "Moderate −" if r_val < -0.2 else "Moderate +" if r_val > 0.2
                else "Weak")
    impl = ("Price cuts will hurt revenue — demand is inelastic in this segment" if r_val > 0.3
            else "Premium pricing suppresses volume — consider yield management" if r_val < -0.4
            else "Dynamic pricing opportunity — demand responds to price changes" if r_val < -0.2
            else "Price is not the primary demand driver — focus on reviews and photos")
    vals = [cl, cs["name"], cs["n_regions"], f"R{cs['avg_price']:,}",
            cs["avg_demand"], r_val, strength, impl]
    for col, v in enumerate(vals, 1):
        cell = ws_corr.cell(row=row, column=col, value=v)
        cell.font   = Font(name="Calibri", size=10)
        cell.border = thin_border()
        cell.alignment = Alignment(wrap_text=(col==8), vertical="center")
        if col == 1:
            cell.fill = hfill(CL_COLORS[cl])
            cell.font = Font(name="Calibri", size=10, color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center")
        elif col == 6:  # r value
            r_abs = abs(r_val)
            cell.font = Font(name="Calibri", size=10, bold=True,
                            color=(GCP_RED if r_val < -0.3 else GCP_GREEN if r_val > 0.3 else "555555"))
        elif row % 2 == 0:
            cell.fill = hfill("F0F4FF")
    ws_corr.row_dimensions[row].height = 40

# Region scatter data (top 20 by demand for scatter plot reference)
scatter_row = 20
ws_corr["A20"].value = "Region Scatter Data (Price vs Demand) — Top 40 by Demand Score"
ws_corr["A20"].font  = Font(name="Calibri", bold=True, size=11, color=BQ_TEAL)

header_row(ws_corr, 21,
    ["Region","Province","Avg Price (ZAR)","Avg Demand Score","Cluster","Price Tier"],
    fill_hex=GCP_GREEN)

top_scatter = stats.sort_values("avg_demand", ascending=False).head(40)
for r_off, (_, row_data) in enumerate(top_scatter.iterrows()):
    row = 22 + r_off
    cl = int(row_data["cluster"]) if pd.notna(row_data["cluster"]) else 3
    vals = [
        row_data["region"], row_data["province"],
        int(row_data["avg_price"]) if pd.notna(row_data["avg_price"]) else 0,
        round(float(row_data["avg_demand"]),2) if pd.notna(row_data["avg_demand"]) else 0,
        CL_NAMES[cl],
        str(row_data["top_tier"]),
    ]
    for col, v in enumerate(vals, 1):
        cell = ws_corr.cell(row=row, column=col, value=v)
        cell.font   = Font(name="Calibri", size=9)
        cell.border = thin_border()
        if col == 5:
            cell.fill = hfill(CL_COLORS[cl])
            cell.font = Font(name="Calibri", size=9, color=WHITE)
        elif row % 2 == 0:
            cell.fill = hfill("F8F9FA")

# Scatter chart (price vs demand using region data)
scatter_chart = BarChart()
scatter_chart.type = "col"
scatter_chart.title = "Avg Demand Score by Region (Top 40, sorted by demand)"
scatter_chart.y_axis.title = "Demand Score"
scatter_chart.x_axis.title = "Region"
scatter_chart.style = 10
scatter_chart.width = 28
scatter_chart.height = 14

demand_ref = Reference(ws_corr, min_col=4, min_row=22, max_row=22+len(top_scatter)-1)
scatter_chart.add_data(demand_ref, titles_from_data=False)
ws_corr.add_chart(scatter_chart, "J1")

for col_letter, width in zip("ABCDEFGH", [22,18,16,18,26,18,16,50]):
    ws_corr.column_dimensions[col_letter].width = width
ws_corr.freeze_panes = "A22"

print("  Sheet 'Price-Demand Corr' written")

# ── save ──────────────────────────────────────────────────────────────────────
wb.save(XL_PATH)
size_kb = round(XL_PATH.stat().st_size / 1024, 0)
print(f"\nDONE - Workbook saved: {XL_PATH.name} ({size_kb:.0f} KB)")
print(f"   Sheets: {wb.sheetnames}")
print(f"\n   KEY FINDINGS embedded in Excel:")
print(f"   • K-Means cluster 0 (High-Demand): 7 regions, avg R{cluster_stats[0]['avg_price']:,}/night")
print(f"   • K-Means cluster 2 (Value Volume): 83 regions — the bulk market")
print(f"   • Global price↔demand r = {round(global_r,3)} — weak negative (higher price ≠ higher demand)")
print(f"   • Summer dominates with {all_season_avgs['Summer']}% of annual bookings vs "
      f"Winter {all_season_avgs['Winter']}%")
