import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

XL_PATH = Path("reports/SA_Accommodation_Intelligence_Platform.xlsx")
wb = openpyxl.load_workbook(XL_PATH)

if "README" in wb.sheetnames:
    del wb["README"]
ws = wb.create_sheet("README", 0)

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 82
ws.row_dimensions[1].height = 45

def hdr(r, c, v, bold=True, sz=11, bg=None, fg="1A1A2E", wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, size=sz, color=fg)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    return cell

hdr(1, 1, "SA ACCOMMODATION INTELLIGENCE PLATFORM  |  LekkeSlaap.co.za Analysis  |  June 2026",
    bold=True, sz=14, bg="1A1A2E", fg="FF6C00", wrap=True)
ws.merge_cells("A1:B1")

rows = [
    ("What is this file?",
     "Excel workbook for the LekkeSlaap Accommodation Intelligence Platform - data analysis of "
     "1,011 listings from LekkeSlaap.co.za (South Africa's leading local accommodation app), "
     "combined with synthetic Google Analytics 4 event data and machine learning insights."),
    ("Data source",
     "LekkeSlaap.co.za scrape - June 2026. 1,011 listings across 114 SA regions. "
     "GA4 events (28,628 rows) are SYNTHETIC - generated to mirror real booking patterns for "
     "demonstration purposes only. All prices are in South African Rand (ZAR / R) per night."),
    ("Currency and units",
     "All monetary values = South African Rand (ZAR / R). Prices = per night. "
     "Demand Score 0-100: higher score = more in-demand relative to platform average. "
     "Review Count = total cumulative guest reviews at scrape date."),
    ("Sheet 2 - Cover",
     "Platform overview, methodology summary, and headline KPIs across all data layers."),
    ("Sheet 3 - Dirty Data Audit",
     "Raw data quality report: nulls, outliers, price inconsistencies found and fixed during ETL. "
     "Shows what was wrong with the source data before cleaning."),
    ("Sheet 4 - Clean Data",
     "1,011 cleaned property records: price_zar, demand_score, review_count, region, "
     "listing_type, price_tier, has_promo_flag, discount_pct, price_outlier_flag."),
    ("Sheet 5 - Regional Analysis",
     "114 regions ranked by avg nightly price, avg demand score, and listing count. "
     "Use this to identify supply gaps and high-opportunity regions."),
    ("Sheet 6 - GA4 Web Analytics",
     "Synthetic GA4 data: session channels, device split (desktop/mobile/tablet), "
     "province distribution, event funnel from page_view to booking_complete."),
    ("Sheet 7 - ML Model Results",
     "5 ML model performance metrics. Random Forest: R2=0.82 (demand prediction). "
     "Gradient Boosting: F1=0.79 (price tier classification). Logistic Regression: AUC=0.74 "
     "(conversion). Isolation Forest: Precision=0.88 (anomaly detection). K-Means: Silhouette=0.61."),
    ("Sheet 8 - ML Predictions",
     "Per-property ML output: predicted demand score, predicted price tier, anomaly_flag "
     "(True = unusual pricing or review pattern flagged by Isolation Forest)."),
    ("Sheet 9 - GTM + GA4 Guide",
     "28 GTM tags documented: tag name, tag type (GA4/Google Ads/Meta Pixel/TikTok/Microsoft UET), "
     "trigger event, POPIA consent mode status. GTM container: GTM-N5J6."),
    ("Sheet 10 - ML Clustering",
     "K-Means clustering (k=4) results by region. "
     "Cluster 0 High-Demand Hotspot: 7 regions, avg R1,297/night, demand 24.04. "
     "Cluster 1 Established Premium: 19 regions, avg R1,651/night, demand 7.17. "
     "Cluster 2 Value Volume Leader: 83 regions, avg R1,851/night, demand 5.64. "
     "Cluster 3 Emerging Gem: 5 regions, avg R9,754/night, demand 0.85."),
    ("Sheet 11 - Seasonal Demand",
     "Booking demand share by SA meteorological season per province. "
     "Summer Dec-Feb = peak beach/Garden Route. Winter Jun-Aug = Kruger/game reserve peak. "
     "Autumn Mar-May = Easter long weekend spike. Spring Sep-Nov = shoulder season. "
     "School holidays add approx 20% demand above seasonal baseline."),
    ("Sheet 12 - Price-Demand Corr",
     "Pearson correlation (r) between nightly price and demand score. "
     "Global r = -0.279: weak negative - cheaper properties get proportionally more bookings. "
     "Cluster 3 Emerging Gems: r = -0.644 (strongest inverse relationship)."),
    ("Demand Score explained",
     "Demand Score = composite: booking-event frequency (40 weight) + review count (30 weight) "
     "+ promotional activity flag (30 weight). Score 50 = platform average. Score 80+ = high-demand. "
     "Score below 20 = low-traffic / emerging property."),
    ("LekkeSlaap app context",
     "LekkeSlaap started in 2013 and is South Africa's leading local accommodation discovery app "
     "(iOS + Android). Primary users are South African domestic travellers. "
     "Mobile attribution tracked via AppsFlyer SDK across Meta Ads, Google UAC, TikTok, and organic channels."),
    ("Reproducibility",
     "Python scripts: 01_etl_clean.py, 02_ml_models.py, 03_generate_outputs.py, "
     "05b_excel_ml_update.py, 06_build_platform.py, 07_appsflyer_seo_audit.py. "
     "BigQuery dataset: africa-south1.accommodation_intelligence. "
     "Scrape date: 2026-06-27."),
]

for i, (label, desc) in enumerate(rows, start=2):
    ws.row_dimensions[i].height = max(32, len(desc) // 3)
    hdr(i, 1, label, bold=True, sz=9, bg="F0F0F0", fg="1A1A2E")
    c = ws.cell(row=i, column=2, value=desc)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(size=9)
    if i % 2 == 0:
        c.fill = PatternFill("solid", fgColor="FAFAFA")

wb.save(XL_PATH)
sz = XL_PATH.stat().st_size // 1024
print(f"Done. {XL_PATH.name}: {sz} KB, {len(wb.sheetnames)} sheets")
print(f"Sheets: {wb.sheetnames}")
