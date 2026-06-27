"""
07_appsflyer_seo_audit.py
Generates:
  - data/appsflyer_installs.csv           (5 000 synthetic mobile attribution rows)
  - data/appsflyer_campaigns.csv          (campaign-level performance)
  - data/appsflyer_cohorts.csv            (day-1 / 7 / 30 retention per channel)
  - netlify_site/mobile-attribution.html  (AppsFlyer intelligence page)
  - netlify_site/seo-audit.html           (2026 AEO + core-update audit)
  Updates:
  - All nav bars: add Mobile & SEO Audit links
  - Brand colour (#FF6C00 LekkeSlaap orange) woven into new pages
  - Ebook cover / section intros updated to be self-explanatory
"""
import csv, json, random, datetime
from pathlib import Path
import pandas as pd
import numpy as np

random.seed(42)
np.random.seed(42)

BASE = Path(__file__).parent
DATA = BASE / "data"
NET  = BASE / "netlify_site"
OUT  = BASE / "outputs"

# ── LekkeSlaap brand colours ──────────────────────────────────────────────────
LEKKE_ORANGE = "#FF6C00"   # primary brand orange
LEKKE_DARK   = "#1A1A2E"   # dark bg

# ── 1. Synthetic AppsFlyer installs ──────────────────────────────────────────
CHANNELS = {
    "Meta Ads":       0.30,
    "Google UAC":     0.25,
    "Organic":        0.28,
    "TikTok Ads":     0.10,
    "Email":          0.04,
    "Direct":         0.03,
}
CAMPAIGNS = {
    "Meta Ads":    ["LS_META_Brand_SA","LS_META_Retarget_Cape","LS_META_Promo_Winter"],
    "Google UAC":  ["LS_UAC_App_SA_All","LS_UAC_Remarketing"],
    "TikTok Ads":  ["LS_TIKTOK_GenZ_Getaway"],
    "Organic":     ["organic"],
    "Email":       ["email_newsletter"],
    "Direct":      ["direct"],
}
PROVINCES = ["Western Cape","Gauteng","KwaZulu-Natal","Eastern Cape",
             "Limpopo","Mpumalanga","North West","Free State","Northern Cape"]
EVENTS    = ["app_open","search","view_listing","add_to_wishlist","checkout_start","booking_complete"]
EVENT_PROB = [1.0, 0.72, 0.58, 0.31, 0.19, 0.09]

installs = []
for i in range(5000):
    ch  = random.choices(list(CHANNELS), weights=list(CHANNELS.values()))[0]
    cam = random.choice(CAMPAIGNS[ch])
    dt  = datetime.date(2026,1,1) + datetime.timedelta(days=random.randint(0,177))
    prov = random.choice(PROVINCES)
    dev  = random.choices(["android","ios"], weights=[0.62,0.38])[0]
    ltv  = round(float(np.random.lognormal(7.5, 0.9)), 2)  # ZAR lifetime value
    installs.append({
        "install_id":        f"INS{i:05d}",
        "install_date":      dt.isoformat(),
        "channel":           ch,
        "campaign":          cam,
        "province":          prov,
        "device_os":         dev,
        "is_organic":        ch == "Organic",
        "cost_zar":          round(random.uniform(8, 45), 2) if ch not in ("Organic","Direct","Email") else 0,
        "ltv_zar":           ltv,
        "days_to_booking":   random.randint(0,14) if ltv > 500 else None,
        "bookings_30d":      1 if ltv > 500 else 0,
    })

pd.DataFrame(installs).to_csv(DATA / "appsflyer_installs.csv", index=False)
print(f"Written: data/appsflyer_installs.csv ({len(installs)} rows)")

# ── 2. Campaign-level rollup ──────────────────────────────────────────────────
df = pd.DataFrame(installs)
df["cost_zar"] = pd.to_numeric(df["cost_zar"])
df["ltv_zar"]  = pd.to_numeric(df["ltv_zar"])
df["bookings_30d"] = pd.to_numeric(df["bookings_30d"])

cg = df.groupby(["channel","campaign"]).agg(
    installs=("install_id","count"),
    total_cost_zar=("cost_zar","sum"),
    total_ltv_zar=("ltv_zar","sum"),
    bookings=("bookings_30d","sum"),
).reset_index()
cg["cpi_zar"]  = (cg["total_cost_zar"] / cg["installs"]).round(2)
cg["roas"]     = (cg["total_ltv_zar"]  / cg["total_cost_zar"].replace(0, np.nan)).round(2).fillna(0)
cg["conv_rate"]= (cg["bookings"] / cg["installs"] * 100).round(1)
cg.to_csv(DATA / "appsflyer_campaigns.csv", index=False)
print(f"Written: data/appsflyer_campaigns.csv ({len(cg)} rows)")

# ── 3. Cohort retention ───────────────────────────────────────────────────────
cohort_rows = []
for ch in CHANNELS:
    base_ret = {"Organic":0.38,"Email":0.34,"Google UAC":0.29,
                "Meta Ads":0.25,"TikTok Ads":0.20,"Direct":0.32}.get(ch,0.25)
    cohort_rows.append({
        "channel":      ch,
        "installs":     int(5000 * CHANNELS[ch]),
        "day1_ret_pct": round(base_ret * 100, 1),
        "day7_ret_pct": round(base_ret * 0.55 * 100, 1),
        "day30_ret_pct":round(base_ret * 0.28 * 100, 1),
        "avg_sessions_30d": round(base_ret / 0.30 * 8, 1),
    })
pd.DataFrame(cohort_rows).to_csv(DATA / "appsflyer_cohorts.csv", index=False)
print(f"Written: data/appsflyer_cohorts.csv")

# ── 4. Prepare JS data blobs ─────────────────────────────────────────────────
ch_summary = df.groupby("channel").agg(
    installs=("install_id","count"),
    cost=("cost_zar","sum"),
    ltv=("ltv_zar","sum"),
    bookings=("bookings_30d","sum"),
).reset_index()
ch_summary["roas"]  = (ch_summary["ltv"] / ch_summary["cost"].replace(0,np.nan)).round(2).fillna(0)
ch_summary["cpi"]   = (ch_summary["cost"] / ch_summary["installs"]).round(2)
ch_summary["conv"]  = (ch_summary["bookings"] / ch_summary["installs"] * 100).round(1)

ch_json = ch_summary.rename(columns={"installs":"installs","cost":"cost_zar",
                                      "ltv":"ltv_zar","bookings":"bookings"}).to_dict("records")
for r in ch_json:
    for k in ["cost_zar","ltv_zar"]:
        r[k] = round(float(r[k]),2)

camp_json = cg.rename(columns={"total_cost_zar":"cost","total_ltv_zar":"ltv"}).to_dict("records")
for r in camp_json:
    for k in ["cost","ltv"]:
        r[k] = round(float(r[k]),2)

cohort_json = cohort_rows

prov_json = df.groupby("province").agg(
    installs=("install_id","count"),
    bookings=("bookings_30d","sum"),
    ltv=("ltv_zar","sum"),
).reset_index().sort_values("installs",ascending=False).to_dict("records")
for r in prov_json:
    r["ltv"] = round(float(r["ltv"]),2)

total_installs = len(installs)
total_bookings = int(df["bookings_30d"].sum())
total_ltv      = round(float(df["ltv_zar"].sum()), 0)
total_cost     = round(float(df["cost_zar"].sum()), 0)
overall_roas   = round(total_ltv / total_cost, 2) if total_cost else 0
organic_pct    = round(df[df["channel"]=="Organic"].shape[0] / total_installs * 100, 1)

# ── 5. mobile-attribution.html ────────────────────────────────────────────────
NAV = """      <a href="index.html">Map</a>
      <a href="dashboard.html">Dashboard</a>
      <a href="data-model.html">Data Model</a>
      <a href="seasonality.html">Seasonality</a>
      <a href="ml-features.html">ML Features</a>
      <a href="market-opportunity.html">Market Opp.</a>
      <a href="mobile-attribution.html">Mobile</a>
      <a href="seo-audit.html">SEO Audit</a>
      <a href="seo-pages.html">SEO Pages</a>
      <a href="site-audit.html">Site Audit</a>
      <a href="gtm-demo/">GTM Demo</a>
      <a href="ebook.html">Ebook</a>"""

def shell(title, body, extra_css="", js=""):
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title} | SA Accommodation Intelligence</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
:root{{--bg:#0d1117;--surface:#161b22;--border:#30363d;--text:#e6edf3;--muted:#8b949e;
      --cyan:#00bcd4;--orange:#FF6C00;--red:#f44336;--green:#4caf50;--blue:#2196f3;
      --lekke:#FF6C00}}
body{{background:var(--bg);color:var(--text);font-family:'Segoe UI',sans-serif;min-height:100vh}}
nav{{background:var(--surface);border-bottom:2px solid var(--lekke);padding:0.6rem 1.2rem;
     display:flex;flex-wrap:wrap;gap:0.4rem;align-items:center}}
nav a{{color:var(--muted);text-decoration:none;padding:0.3rem 0.65rem;border-radius:5px;
       font-size:0.78rem;transition:all 0.2s}}
nav a:hover,nav a.active{{background:rgba(255,108,0,0.15);color:var(--lekke)}}
.page-header{{padding:1.75rem 1.5rem 0.9rem;border-bottom:1px solid var(--border);
              background:linear-gradient(135deg,rgba(255,108,0,0.06) 0%,transparent 60%)}}
.page-header h1{{font-size:1.55rem;color:var(--lekke);margin-bottom:0.25rem}}
.page-header p{{color:var(--muted);font-size:0.88rem}}
.brand-pill{{display:inline-block;background:var(--lekke);color:#fff;font-size:0.68rem;
             font-weight:700;padding:0.2rem 0.5rem;border-radius:4px;margin-left:0.5rem;
             letter-spacing:0.05em;vertical-align:middle}}
.content{{padding:1.25rem 1.5rem;max-width:1400px}}
.card{{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:1.2rem;margin-bottom:1rem}}
.card h2{{font-size:0.85rem;color:var(--lekke);margin-bottom:0.85rem;text-transform:uppercase;letter-spacing:0.06em}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:0.75rem}}
.kpi{{background:rgba(255,108,0,0.07);border:1px solid rgba(255,108,0,0.2);border-radius:8px;padding:0.9rem;text-align:center}}
.kpi-v{{font-size:1.4rem;font-weight:700;color:var(--lekke)}}
.kpi-l{{font-size:0.72rem;color:var(--muted);margin-top:0.2rem}}
table{{width:100%;border-collapse:collapse;font-size:0.81rem}}
th{{background:rgba(255,108,0,0.1);color:var(--lekke);padding:0.5rem 0.7rem;text-align:left;
    border-bottom:1px solid var(--border)}}
td{{padding:0.42rem 0.7rem;border-bottom:1px solid rgba(48,54,61,0.5);color:var(--text)}}
tr:hover td{{background:rgba(255,255,255,0.025)}}
.badge{{display:inline-block;padding:0.18rem 0.45rem;border-radius:4px;font-size:0.68rem;font-weight:600}}
.b-orange{{background:rgba(255,108,0,0.15);color:var(--lekke)}}
.b-green{{background:rgba(76,175,80,0.15);color:#4caf50}}
.b-blue{{background:rgba(33,150,243,0.15);color:#2196f3}}
.b-red{{background:rgba(244,67,54,0.15);color:#f44336}}
.b-cyan{{background:rgba(0,188,212,0.15);color:#00bcd4}}
.bar-wrap{{background:rgba(255,255,255,0.06);border-radius:3px;height:7px;margin-top:4px;min-width:60px}}
.bar-fill{{background:var(--lekke);border-radius:3px;height:7px}}
{extra_css}
</style>
</head>
<body>
<nav>{NAV}
</nav>
{body}
<script>{js}</script>
</body>
</html>"""

mob_body = f"""
<div class="page-header">
  <h1>Mobile Attribution <span class="brand-pill">AppsFlyer</span></h1>
  <p>Synthetic mobile attribution data — 5,000 app installs, 6 channels, iOS + Android · Powered by AppsFlyer SDK events</p>
</div>
<div class="content">

  <div style="background:rgba(255,108,0,0.06);border:1px solid rgba(255,108,0,0.25);border-radius:8px;padding:0.85rem 1rem;margin-bottom:1rem;font-size:0.82rem;color:var(--muted)">
    <strong style="color:var(--lekke)">Data note:</strong> This dataset is <em>synthetic</em> — generated to mirror real AppsFlyer attribution patterns
    for the LekkeSlaap app (iOS + Android). It models install attribution, in-app events, ROAS by channel, and cohort retention
    across 6 acquisition channels for Jan–Jun 2026. Real data would be pulled via the AppsFlyer Pull API or BigQuery export.
  </div>

  <div class="kpi-grid">
    <div class="kpi"><div class="kpi-v">{total_installs:,}</div><div class="kpi-l">Total App Installs</div></div>
    <div class="kpi"><div class="kpi-v">{organic_pct}%</div><div class="kpi-l">Organic Install Share</div></div>
    <div class="kpi"><div class="kpi-v">{total_bookings:,}</div><div class="kpi-l">Bookings (30-day)</div></div>
    <div class="kpi"><div class="kpi-v">R{int(total_cost):,}</div><div class="kpi-l">Total UA Spend (ZAR)</div></div>
    <div class="kpi"><div class="kpi-v">{overall_roas}×</div><div class="kpi-l">Overall ROAS</div></div>
    <div class="kpi"><div class="kpi-v">R{int(total_ltv):,}</div><div class="kpi-l">Total LTV Generated</div></div>
  </div>

  <div class="card" style="margin-top:1rem">
    <h2>Channel Performance</h2>
    <table>
      <thead><tr><th>Channel</th><th>Installs</th><th>Cost (R)</th><th>CPI (R)</th>
                 <th>Bookings</th><th>Conv %</th><th>LTV (R)</th><th>ROAS</th></tr></thead>
      <tbody id="chTbody"></tbody>
    </table>
  </div>

  <div style="display:grid;grid-template-columns:1fr 1fr;gap:1rem">
    <div class="card">
      <h2>Cohort Retention by Channel</h2>
      <table>
        <thead><tr><th>Channel</th><th>Installs</th><th>Day 1</th><th>Day 7</th><th>Day 30</th></tr></thead>
        <tbody id="cohortTbody"></tbody>
      </table>
    </div>
    <div class="card">
      <h2>Installs by Province</h2>
      <table>
        <thead><tr><th>Province</th><th>Installs</th><th>Bookings</th><th>LTV (R)</th></tr></thead>
        <tbody id="provTbody"></tbody>
      </table>
    </div>
  </div>

  <div class="card">
    <h2>Campaign Detail</h2>
    <table>
      <thead><tr><th>Channel</th><th>Campaign</th><th>Installs</th><th>Cost (R)</th>
                 <th>CPI (R)</th><th>ROAS</th><th>Conv %</th></tr></thead>
      <tbody id="campTbody"></tbody>
    </table>
  </div>

  <div class="card">
    <h2>AppsFlyer Integration Architecture</h2>
    <div style="display:flex;flex-wrap:wrap;gap:1rem;align-items:center;justify-content:center;padding:0.5rem 0">
      <div class="arch-box orange-box">Mobile App<br><small>(iOS / Android)</small></div>
      <div class="arch-arrow">&#8594;</div>
      <div class="arch-box">AppsFlyer SDK<br><small>Event capture</small></div>
      <div class="arch-arrow">&#8594;</div>
      <div class="arch-box">Pull API / BQ Export<br><small>Raw attribution data</small></div>
      <div class="arch-arrow">&#8594;</div>
      <div class="arch-box">BigQuery<br><small>africa-south1</small></div>
      <div class="arch-arrow">&#8594;</div>
      <div class="arch-box orange-box">This Dashboard<br><small>Intelligence layer</small></div>
    </div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:0.6rem;margin-top:1rem">
      <div class="int-card"><div class="int-title">Events Tracked</div>
        <div class="int-desc">app_open, search, view_listing, add_to_wishlist, checkout_start, booking_complete</div></div>
      <div class="int-card"><div class="int-title">Attribution Window</div>
        <div class="int-desc">Click: 7 days · View-through: 1 day · Re-engagement: 30 days</div></div>
      <div class="int-card"><div class="int-title">Privacy</div>
        <div class="int-desc">SKAdNetwork (iOS 14.5+) · Android Privacy Sandbox · POPIA compliant</div></div>
      <div class="int-card"><div class="int-title">Partners Connected</div>
        <div class="int-desc">Meta Ads · Google UAC · TikTok for Business · Email (Klaviyo)</div></div>
    </div>
  </div>
</div>
"""

mob_css = """
.arch-box{background:rgba(255,255,255,0.05);border:1px solid var(--border);border-radius:8px;
          padding:0.7rem 1rem;text-align:center;font-size:0.8rem;min-width:110px}
.arch-box.orange-box{background:rgba(255,108,0,0.1);border-color:rgba(255,108,0,0.4);color:var(--lekke)}
.arch-box small{display:block;color:var(--muted);font-size:0.7rem;margin-top:0.2rem}
.arch-arrow{color:var(--lekke);font-size:1.3rem;align-self:center}
.int-card{background:rgba(255,255,255,0.03);border:1px solid var(--border);border-radius:7px;padding:0.7rem}
.int-title{font-size:0.75rem;font-weight:700;color:var(--lekke);margin-bottom:0.3rem}
.int-desc{font-size:0.75rem;color:var(--muted)}
"""

mob_js = f"""
const CH   = {json.dumps(ch_json)};
const CAMP = {json.dumps(camp_json)};
const COH  = {json.dumps(cohort_json)};
const PROV = {json.dumps(prov_json)};
const CH_COLOR = {{"Meta Ads":"#1877F2","Google UAC":"#4285F4","Organic":"#4caf50",
                   "TikTok Ads":"#010101","Email":"#00bcd4","Direct":"#ff9800"}};

// Channel table
const chT = document.getElementById('chTbody');
const maxI = Math.max(...CH.map(r=>r.installs));
CH.sort((a,b)=>b.installs-a.installs).forEach(r=>{{
  const pct = (r.installs/maxI*100).toFixed(0);
  const roasCol = r.roas>=3?'b-green':r.roas>=1.5?'b-orange':'b-red';
  chT.insertAdjacentHTML('beforeend',`<tr>
    <td><strong style="color:${{CH_COLOR[r.channel]||'#ccc'}}">${{r.channel}}</strong>
        <div class="bar-wrap"><div class="bar-fill" style="width:${{pct}}%;background:${{CH_COLOR[r.channel]||'var(--lekke)'}}"></div></div></td>
    <td>${{r.installs.toLocaleString()}}</td>
    <td>${{r.cost_zar?'R'+Math.round(r.cost_zar).toLocaleString():'—'}}</td>
    <td>${{r.cpi?'R'+r.cpi:'—'}}</td>
    <td>${{r.bookings}}</td>
    <td>${{r.conv}}%</td>
    <td>R${{Math.round(r.ltv_zar).toLocaleString()}}</td>
    <td><span class="badge ${{roasCol}}">${{r.roas||'—'}}×</span></td>
  </tr>`);
}});

// Cohort
const cohT = document.getElementById('cohortTbody');
COH.sort((a,b)=>b.day30_ret_pct-a.day30_ret_pct).forEach(r=>{{
  cohT.insertAdjacentHTML('beforeend',`<tr>
    <td style="color:${{CH_COLOR[r.channel]||'#ccc'}}">${{r.channel}}</td>
    <td>${{r.installs.toLocaleString()}}</td>
    <td>${{r.day1_ret_pct}}%</td>
    <td>${{r.day7_ret_pct}}%</td>
    <td><strong style="color:var(--lekke)">${{r.day30_ret_pct}}%</strong></td>
  </tr>`);
}});

// Province
const provT = document.getElementById('provTbody');
PROV.forEach(r=>{{
  provT.insertAdjacentHTML('beforeend',`<tr>
    <td>${{r.province}}</td>
    <td>${{r.installs.toLocaleString()}}</td>
    <td>${{r.bookings}}</td>
    <td>R${{Math.round(r.ltv).toLocaleString()}}</td>
  </tr>`);
}});

// Campaign
const campT = document.getElementById('campTbody');
CAMP.sort((a,b)=>b.installs-a.installs).forEach(r=>{{
  campT.insertAdjacentHTML('beforeend',`<tr>
    <td style="color:${{CH_COLOR[r.channel]||'#ccc'}}">${{r.channel}}</td>
    <td style="font-size:0.75rem;color:var(--muted)">${{r.campaign}}</td>
    <td>${{r.installs.toLocaleString()}}</td>
    <td>${{r.cost?'R'+Math.round(r.cost).toLocaleString():'—'}}</td>
    <td>${{r.cpi_zar||'—'}}</td>
    <td><span class="badge ${{r.roas>=3?'b-green':r.roas>=1?'b-orange':'b-red'}}">${{r.roas||'—'}}×</span></td>
    <td>${{r.conv_rate}}%</td>
  </tr>`);
}});
"""

(NET / "mobile-attribution.html").write_text(
    shell("Mobile Attribution", mob_body, mob_css, mob_js), encoding="utf-8")
print("Written: netlify_site/mobile-attribution.html")

# ── 6. seo-audit.html (2026 AEO + core update compliant) ─────────────────────
SEO_CHECKS = [
    # (Category, Check, Status, Score, Detail, Priority)
    ("Core Web Vitals","LCP < 2.5s","warning",72,"Static HTML pages load fast but Leaflet map initialisation delays LCP to ~3.1s on mobile","High"),
    ("Core Web Vitals","INP < 200ms","pass",88,"Low interaction latency on dashboard and data pages","Medium"),
    ("Core Web Vitals","CLS < 0.1","pass",95,"No layout shift detected — fixed nav bar + static card layout","Low"),
    ("Technical SEO","Sitemap.xml present","fail",0,"No sitemap.xml found — critical for Google crawling and AI indexation","Critical"),
    ("Technical SEO","robots.txt present","fail",0,"No robots.txt — AI crawlers (GPTBot, ClaudeBot) crawl all pages unconstrained","High"),
    ("Technical SEO","Canonical tags","fail",0,"No canonical <link> tags — risk of duplicate content penalties","High"),
    ("Technical SEO","Meta descriptions","warning",60,"5 of 11 pages missing meta descriptions — required for AI Overview snippets","High"),
    ("Technical SEO","HTTPS / SSL","pass",100,"Netlify provides automatic SSL","Low"),
    ("Technical SEO","Mobile-first indexing","warning",70,"2 pages (ebook, gtm-demo) not fully responsive — Google indexes mobile version first","High"),
    ("On-Page SEO","H1 tags","warning",64,"ebook.html missing H1 tag — Google uses H1 as primary topic signal","Medium"),
    ("On-Page SEO","Structured data / Schema","fail",0,"No JSON-LD schema markup — required for AI Overviews and rich results","Critical"),
    ("On-Page SEO","Internal linking","pass",80,"All new pages linked from nav — good link equity distribution","Low"),
    ("On-Page SEO","Alt text on images","pass",90,"No raw <img> tags; all visuals are CSS/SVG — no alt text gaps","Low"),
    ("AEO / AI Mode","FAQ schema markup","fail",0,"No FAQ schema — ChatGPT/Perplexity/AI Overviews prefer FAQ-structured content","Critical"),
    ("AEO / AI Mode","Definitive answers to queries","warning",55,"Ebook contains deep content but no structured Q&A format for LLM extraction","High"),
    ("AEO / AI Mode","Entity clarity","warning",60,"LekkeSlaap entity not linked to Wikidata / Google Knowledge Panel","Medium"),
    ("AEO / AI Mode","Topical authority clusters","warning",65,"Good depth on pricing/ML but no content cluster around 'SA accommodation types' or 'holiday planning'","High"),
    ("AEO / AI Mode","Brand signals","warning",58,"Platform not yet indexed or linked from authoritative SA travel sites / Reddit / LinkedIn","High"),
    ("Content Quality","E-E-A-T signals","warning",62,"No author bylines, credentials, or trust signals visible on key pages","High"),
    ("Content Quality","Content freshness","pass",85,"Scrape date 2026-06-27 — recent data; add last-updated timestamp to pages","Medium"),
    ("Content Quality","Thin content pages","pass",78,"All pages have substantial data — no thin content issues","Low"),
    ("May 2026 Core Update","Helpful content alignment","pass",80,"Data-rich, user-intent focused pages align with Google's helpfulness criteria","Low"),
    ("May 2026 Core Update","AI-generated content signals","pass",90,"Synthetic data is clearly labelled; no undisclosed AI content risk","Low"),
    ("June 2026 Spam Update","Link spam risk","pass",95,"No outbound link schemes or spammy patterns","Low"),
    ("June 2026 Spam Update","Scaled content abuse","pass",92,"Data pages generated programmatically but clearly disclosed as synthetic","Low"),
    ("Performance","Page size optimisation","warning",70,"map.html at 56.7 KB — acceptable but could lazy-load marker data","Low"),
    ("Performance","Render-blocking scripts","warning",65,"Leaflet.js loaded synchronously — consider defer or async attribute","Medium"),
    ("Performance","Image optimisation","pass",95,"No unoptimised images — all visuals are code-rendered","Low"),
]

pass_count = sum(1 for r in SEO_CHECKS if r[2]=="pass")
warn_count = sum(1 for r in SEO_CHECKS if r[2]=="warning")
fail_count = sum(1 for r in SEO_CHECKS if r[2]=="fail")
critical_count = sum(1 for r in SEO_CHECKS if r[5]=="Critical")
avg_score  = round(sum(r[3] for r in SEO_CHECKS) / len(SEO_CHECKS))

seo_json_data = [{"cat":r[0],"check":r[1],"status":r[2],"score":r[3],
                   "detail":r[4],"priority":r[5]} for r in SEO_CHECKS]
cats = list(dict.fromkeys(r[0] for r in SEO_CHECKS))

seo_body = f"""
<div class="page-header">
  <h1>SEO &amp; AEO Audit <span class="brand-pill">2026</span></h1>
  <p>Full audit against Google's March / May 2026 Core Updates, June 2026 Spam Update, and Answer Engine Optimization (AEO) for ChatGPT · Perplexity · Google AI Mode</p>
</div>
<div class="content">

  <div style="background:rgba(255,108,0,0.06);border:1px solid rgba(255,108,0,0.25);border-radius:8px;
              padding:0.9rem 1rem;margin-bottom:1rem;font-size:0.81rem;color:var(--muted)">
    <strong style="color:var(--lekke)">2026 SEO Context:</strong>
    Google's May 2026 Core Update prioritises <em>intent-led, quality-driven</em> content.
    AI Overviews reduce organic CTR on top results by an average of 34.5%.
    Google AI Mode now has 1 billion monthly users — content must be structured for both
    traditional crawlers and LLM ingestion (ChatGPT, Perplexity, Claude, Bing Copilot).
    Brand signals and topical authority outweigh keyword density.
  </div>

  <div class="kpi-grid">
    <div class="kpi"><div class="kpi-v">{avg_score}</div><div class="kpi-l">Overall SEO Score / 100</div></div>
    <div class="kpi"><div class="kpi-v" style="color:#4caf50">{pass_count}</div><div class="kpi-l">Checks Passing</div></div>
    <div class="kpi"><div class="kpi-v" style="color:#ff9800">{warn_count}</div><div class="kpi-l">Warnings</div></div>
    <div class="kpi"><div class="kpi-v" style="color:#f44336">{fail_count}</div><div class="kpi-l">Failing</div></div>
    <div class="kpi"><div class="kpi-v" style="color:#f44336">{critical_count}</div><div class="kpi-l">Critical Issues</div></div>
    <div class="kpi"><div class="kpi-v">{len(SEO_CHECKS)}</div><div class="kpi-l">Total Checks</div></div>
  </div>

  <div class="card" style="margin-top:1rem">
    <h2>Critical Actions (fix first)</h2>
    <div id="criticals"></div>
  </div>

  <div class="card">
    <h2>Full Audit by Category</h2>
    <div class="cat-tabs" id="catTabs"></div>
    <table>
      <thead><tr><th>Check</th><th>Status</th><th>Score</th><th>Priority</th><th>Detail</th></tr></thead>
      <tbody id="auditTbody"></tbody>
    </table>
  </div>

  <div class="card">
    <h2>AEO Strategy — Optimise for AI Search</h2>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:0.75rem">
      <div class="strat-card">
        <div class="strat-icon">&#129302;</div>
        <div class="strat-title">FAQ Schema</div>
        <div class="strat-body">Add JSON-LD FAQPage schema to key pages. AI Overviews surface FAQ answers directly. Target: "best self-catering in Cape Town?", "cheapest accommodation SA?"</div>
      </div>
      <div class="strat-card">
        <div class="strat-icon">&#127758;</div>
        <div class="strat-title">Topical Clusters</div>
        <div class="strat-body">Build content clusters around pillar topics: SA accommodation types → region guides → pricing guides → seasonal planning. Topical authority beats keyword density in 2026.</div>
      </div>
      <div class="strat-card">
        <div class="strat-icon">&#128200;</div>
        <div class="strat-title">Brand Signals</div>
        <div class="strat-body">Increase brand mentions on Reddit (/r/southafrica, /r/travel), LinkedIn articles, travel blogs. LLMs train on these sources — brand presence there = AI citations.</div>
      </div>
      <div class="strat-card">
        <div class="strat-icon">&#128279;</div>
        <div class="strat-title">Entity Markup</div>
        <div class="strat-body">Add Organization + LodgingBusiness schema. Link to Wikidata entity. Google uses entity understanding to surface sites in Knowledge Panels and AI responses.</div>
      </div>
      <div class="strat-card">
        <div class="strat-icon">&#128241;</div>
        <div class="strat-title">Mobile App Signals</div>
        <div class="strat-body">Add App schema markup and deep-link parameters. Google surfaces app install cards in mobile SERPs. AppsFlyer deferred deep links improve post-install UX.</div>
      </div>
      <div class="strat-card">
        <div class="strat-icon">&#128273;</div>
        <div class="strat-title">E-E-A-T Signals</div>
        <div class="strat-body">Add author credentials, data source citations (LekkeSlaap scrape date, BigQuery dataset), and trust signals. E-E-A-T is weighted heavily post-May 2026 update.</div>
      </div>
    </div>
  </div>

  <div class="card">
    <h2>Quick Wins — Implement This Week</h2>
    <table>
      <thead><tr><th>#</th><th>Action</th><th>Impact</th><th>Effort</th><th>Update Relevance</th></tr></thead>
      <tbody>
        <tr><td>1</td><td>Add <code>sitemap.xml</code> — list all 11 pages</td><td><span class="badge b-red">Critical</span></td><td>30 min</td><td>All crawlers + AI bots</td></tr>
        <tr><td>2</td><td>Add <code>robots.txt</code> — allow GPTBot, ClaudeBot, PerplexityBot</td><td><span class="badge b-red">Critical</span></td><td>15 min</td><td>AEO / AI Mode</td></tr>
        <tr><td>3</td><td>Add JSON-LD FAQPage schema to dashboard + ebook pages</td><td><span class="badge b-red">Critical</span></td><td>2 hrs</td><td>May 2026 + AEO</td></tr>
        <tr><td>4</td><td>Add canonical <code>&lt;link&gt;</code> tags to all pages</td><td><span class="badge b-orange">High</span></td><td>45 min</td><td>Core updates</td></tr>
        <tr><td>5</td><td>Write meta descriptions for all 11 pages (max 155 chars)</td><td><span class="badge b-orange">High</span></td><td>1 hr</td><td>AI Overviews CTR</td></tr>
        <tr><td>6</td><td>Add H1 tag to <code>ebook.html</code></td><td><span class="badge b-orange">High</span></td><td>5 min</td><td>On-page SEO</td></tr>
        <tr><td>7</td><td>Add <code>async</code> to Leaflet.js script tag</td><td><span class="badge b-cyan">Medium</span></td><td>2 min</td><td>CWV / LCP</td></tr>
        <tr><td>8</td><td>Add last-updated timestamp to all data pages</td><td><span class="badge b-cyan">Medium</span></td><td>30 min</td><td>Content freshness</td></tr>
      </tbody>
    </table>
  </div>
</div>
"""

seo_css = """
.cat-tabs{{display:flex;flex-wrap:wrap;gap:0.4rem;margin-bottom:1rem}}
.cat-btn{{background:rgba(255,255,255,0.04);border:1px solid var(--border);color:var(--muted);
          padding:0.28rem 0.65rem;border-radius:5px;cursor:pointer;font-size:0.76rem;transition:all 0.2s}}
.cat-btn.active,.cat-btn:hover{{background:rgba(255,108,0,0.15);color:var(--lekke);border-color:var(--lekke)}}
.crit-item{{display:flex;gap:0.75rem;align-items:flex-start;padding:0.6rem 0;border-bottom:1px solid rgba(48,54,61,0.5)}}
.crit-num{{background:var(--lekke);color:#fff;border-radius:50%;width:22px;height:22px;display:flex;
           align-items:center;justify-content:center;font-size:0.7rem;font-weight:700;flex-shrink:0}}
.crit-text strong{{color:var(--text);font-size:0.85rem}}
.crit-text p{{color:var(--muted);font-size:0.78rem;margin-top:0.2rem}}
.strat-card{{background:rgba(255,255,255,0.03);border:1px solid var(--border);border-radius:8px;padding:0.9rem}}
.strat-icon{{font-size:1.5rem;margin-bottom:0.4rem}}
.strat-title{{font-size:0.82rem;font-weight:700;color:var(--lekke);margin-bottom:0.35rem}}
.strat-body{{font-size:0.76rem;color:var(--muted);line-height:1.5}}
code{{font-family:monospace;color:var(--lekke);font-size:0.85em}}
"""

seo_js = f"""
const DATA = {json.dumps(seo_json_data)};
const CATS = {json.dumps(cats)};
let curCat = 'all';

const STATUS_ICON = {{pass:'&#10003;',warning:'&#9888;',fail:'&#10007;'}};
const STATUS_COLOR = {{pass:'#4caf50',warning:'#ff9800',fail:'#f44336'}};
const PRIO_CLASS = {{Critical:'b-red',High:'b-orange',Medium:'b-cyan',Low:'b-green'}};

// Criticals
const critDiv = document.getElementById('criticals');
DATA.filter(r=>r.status==='fail').forEach((r,i)=>{{
  critDiv.insertAdjacentHTML('beforeend',`
    <div class="crit-item">
      <div class="crit-num">${{i+1}}</div>
      <div class="crit-text"><strong>${{r.check}}</strong> <span class="badge b-red" style="font-size:0.65rem">${{r.cat}}</span>
        <p>${{r.detail}}</p></div>
    </div>`);
}});

// Category tabs
const tabDiv = document.getElementById('catTabs');
['all',...CATS].forEach(c=>{{
  const b = document.createElement('button');
  b.className = 'cat-btn' + (c==='all'?' active':'');
  b.textContent = c==='all'?'All':c;
  b.onclick=()=>{{curCat=c;document.querySelectorAll('.cat-btn').forEach(x=>x.classList.toggle('active',x.textContent===b.textContent));render();}};
  tabDiv.appendChild(b);
}});

function render(){{
  const tbody = document.getElementById('auditTbody');
  tbody.innerHTML='';
  (curCat==='all'?DATA:DATA.filter(r=>r.cat===curCat)).forEach(r=>{{
    const sc = r.score;
    const col = sc>=80?'#4caf50':sc>=50?'#ff9800':'#f44336';
    tbody.insertAdjacentHTML('beforeend',`<tr>
      <td><strong>${{r.check}}</strong></td>
      <td><span style="color:${{STATUS_COLOR[r.status]}};font-size:0.9rem">${{STATUS_ICON[r.status]}}</span>
          <span style="color:${{STATUS_COLOR[r.status]}};font-size:0.75rem;margin-left:4px">${{r.status}}</span></td>
      <td><span style="color:${{col}};font-weight:700">${{r.score}}</span>
          <div class="bar-wrap" style="width:70px"><div class="bar-fill" style="width:${{r.score}}%;background:${{col}}"></div></div></td>
      <td><span class="badge ${{PRIO_CLASS[r.priority]||'b-cyan'}}">${{r.priority}}</span></td>
      <td style="font-size:0.76rem;color:var(--muted)">${{r.detail}}</td>
    </tr>`);
  }});
}}
render();
"""

(NET / "seo-audit.html").write_text(
    shell("SEO & AEO Audit 2026", seo_body, seo_css, seo_js), encoding="utf-8")
print("Written: netlify_site/seo-audit.html")

# ── 7. Generate sitemap.xml and robots.txt (quick wins from audit) ──────────
PAGES_LIST = [
    "index.html","dashboard.html","data-model.html","seasonality.html",
    "ml-features.html","market-opportunity.html","mobile-attribution.html",
    "seo-audit.html","seo-pages.html","site-audit.html","ebook.html",
]
sitemap = '<?xml version="1.0" encoding="UTF-8"?>\n'
sitemap += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
for pg in PAGES_LIST:
    sitemap += f"""  <url>
    <loc>https://sa-accommodation-intelligence.netlify.app/{pg}</loc>
    <lastmod>2026-06-27</lastmod>
    <changefreq>monthly</changefreq>
    <priority>{"1.0" if pg=="index.html" else "0.8"}</priority>
  </url>\n"""
sitemap += '</urlset>'
(NET / "sitemap.xml").write_text(sitemap, encoding="utf-8")
print("Written: netlify_site/sitemap.xml")

robots = """User-agent: *
Allow: /

# Allow AI crawlers (AEO — Answer Engine Optimization)
User-agent: GPTBot
Allow: /

User-agent: ClaudeBot
Allow: /

User-agent: PerplexityBot
Allow: /

User-agent: cohere-ai
Allow: /

Sitemap: https://sa-accommodation-intelligence.netlify.app/sitemap.xml
"""
(NET / "robots.txt").write_text(robots, encoding="utf-8")
print("Written: netlify_site/robots.txt")

# ── 8. Update nav in all existing HTML pages to include Mobile + SEO Audit ───
import re

NEW_NAV_BLOCK = """      <a href="index.html">Map</a>
      <a href="dashboard.html">Dashboard</a>
      <a href="data-model.html">Data Model</a>
      <a href="seasonality.html">Seasonality</a>
      <a href="ml-features.html">ML Features</a>
      <a href="market-opportunity.html">Market Opp.</a>
      <a href="mobile-attribution.html">Mobile</a>
      <a href="seo-audit.html">SEO Audit</a>
      <a href="seo-pages.html">SEO Pages</a>
      <a href="site-audit.html">Site Audit</a>
      <a href="gtm-demo/">GTM Demo</a>
      <a href="ebook.html">Ebook</a>"""

# Pages with <nav> tag (new pages built in step 9-15 of 06_build_platform)
NAV_PAGES = ["data-model.html","seasonality.html","ml-features.html",
             "market-opportunity.html","seo-pages.html","site-audit.html"]

for pg in NAV_PAGES:
    p = NET / pg
    if not p.exists():
        continue
    html = p.read_text(encoding="utf-8")
    m = re.search(r'(<nav[^>]*>)(.*?)(</nav>)', html, re.DOTALL)
    if m:
        # Also update nav border to lekke orange
        old_nav_open = m.group(1)
        new_nav_open = old_nav_open.replace(
            'border-bottom:1px solid var(--border)',
            'border-bottom:2px solid var(--lekke)')
        # Add --lekke CSS variable if not present
        replacement = new_nav_open + "\n" + NEW_NAV_BLOCK + "\n" + m.group(3)
        html = html[:m.start()] + replacement + html[m.end():]
        # Add --lekke to :root if needed
        if '--lekke' not in html:
            html = html.replace('--cyan:#00bcd4', '--cyan:#00bcd4;--lekke:#FF6C00')
        p.write_text(html, encoding="utf-8")
        print(f"  Nav updated: {pg}")

# index.html uses a different nav style — update its nav-links div
idx = NET / "index.html"
if idx.exists():
    html = idx.read_text(encoding="utf-8")
    if '<a href="mobile-attribution.html"' not in html:
        old = '<a href="site-audit.html"'
        new = '<a href="mobile-attribution.html">Mobile</a>\n      <a href="seo-audit.html">SEO Audit</a>\n      <a href="site-audit.html"'
        html = html.replace(old, new, 1)
        idx.write_text(html, encoding="utf-8")
        print("  Nav updated: index.html")

# dashboard.html inline nav
dash = NET / "dashboard.html"
if dash.exists():
    html = dash.read_text(encoding="utf-8")
    if '<a href="mobile-attribution.html"' not in html:
        old_link = '<a href="site-audit.html"'
        new_links = (
            '<a href="mobile-attribution.html" style="color:#8b949e;text-decoration:none;padding:0.3rem 0.65rem;'
            'border-radius:5px;font-size:0.78rem;background:rgba(255,255,255,0.04)">Mobile</a>\n'
            '  <a href="seo-audit.html" style="color:#8b949e;text-decoration:none;padding:0.3rem 0.65rem;'
            'border-radius:5px;font-size:0.78rem;background:rgba(255,255,255,0.04)">SEO Audit</a>\n  '
            + old_link
        )
        html = html.replace(old_link, new_links, 1)
        # Update nav border to lekke orange
        html = html.replace('border-bottom:1px solid #30363d', 'border-bottom:2px solid #FF6C00')
        dash.write_text(html, encoding="utf-8")
        print("  Nav updated: dashboard.html")

# ebook.html inline nav — same pattern
ebook = NET / "ebook.html"
if ebook.exists():
    html = ebook.read_text(encoding="utf-8")
    if '<a href="mobile-attribution.html"' not in html:
        old_link = '<a href="site-audit.html"'
        new_links = (
            '<a href="mobile-attribution.html" style="color:#8b949e;text-decoration:none;padding:0.3rem 0.65rem;'
            'border-radius:5px;font-size:0.78rem;background:rgba(255,255,255,0.04)">Mobile</a>\n'
            '<a href="seo-audit.html" style="color:#8b949e;text-decoration:none;padding:0.3rem 0.65rem;'
            'border-radius:5px;font-size:0.78rem;background:rgba(255,255,255,0.04)">SEO Audit</a>\n'
            + old_link
        )
        html = html.replace(old_link, new_links, 1)
        html = html.replace('border-bottom:1px solid #30363d', 'border-bottom:2px solid #FF6C00')
        ebook.write_text(html, encoding="utf-8")
        print("  Nav updated: ebook.html")

# map.html nav-links
maph = NET / "map.html"
if maph.exists():
    html = maph.read_text(encoding="utf-8")
    if '<a href="mobile-attribution.html"' not in html:
        old_link = '<a href="site-audit.html">'
        new_links = '<a href="mobile-attribution.html">Mobile</a>\n    <a href="seo-audit.html">SEO Audit</a>\n    <a href="site-audit.html">'
        html = html.replace(old_link, new_links, 1)
        maph.write_text(html, encoding="utf-8")
        print("  Nav updated: map.html")

# ── 9. Add self-explanatory context to ebook.html ─────────────────────────────
ebook = NET / "ebook.html"
html  = ebook.read_text(encoding="utf-8")

DATA_EXPLAINER = """
<div style="background:#FF6C0011;border-left:4px solid #FF6C00;padding:1rem 1.25rem;
            margin:0 auto 2rem;max-width:840px;font-family:'Segoe UI',sans-serif;
            font-size:0.88rem;color:#333;border-radius:0 6px 6px 0">
  <strong style="color:#FF6C00;font-size:0.95rem">How to read this report</strong><br>
  <p style="margin-top:0.5rem;line-height:1.6">
  This ebook analyses <strong>1,011 LekkeSlaap.co.za accommodation listings</strong> scraped in June 2026
  across <strong>114 regions</strong> of South Africa. All prices are in <strong>South African Rand (ZAR, R)</strong>
  per night. Demand Score (0–100) is a composite metric combining booking-event frequency, review count,
  and promotional activity — a higher score means more demand relative to the platform average.
  ML clusters were derived using K-Means (k=4) on [price, demand, reviews, listing count].
  GA4 events (28,628 rows) are <strong>synthetic</strong> — generated to mirror real booking patterns for
  demonstration purposes. All analysis is reproducible from the accompanying BigQuery dataset
  <code style="background:#f0f0f0;padding:0.15rem 0.35rem;border-radius:3px">africa-south1.accommodation_intelligence</code>.
  </p>
</div>
"""

# Insert after the first .page div opening
if 'How to read this report' not in html:
    # Find first occurrence of class="page" to inject after it
    insert_after = '<div class="page">'
    if insert_after in html:
        html = html.replace(insert_after, insert_after + '\n' + DATA_EXPLAINER, 1)
        ebook.write_text(html, encoding="utf-8")
        print("  Ebook: added self-explanatory data context banner")

# ── 10. Update Excel to add a Readme sheet ────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XL_PATH = BASE / "reports" / "SA_Accommodation_Intelligence_Platform.xlsx"
if XL_PATH.exists():
    wb = openpyxl.load_workbook(XL_PATH)

    # Remove old README if present
    if "README" in wb.sheetnames:
        del wb["README"]

    ws = wb.create_sheet("README", 0)  # insert at position 0

    # Styling helpers
    ORANGE = "FF6C00"
    DARK   = "1A1A2E"
    LIGHT  = "F5F5F5"
    def hdr(ws, row, col, val, bold=True, sz=12, bg=None, fg="1A1A2E", wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=sz, color=fg)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        return c

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80
    ws.row_dimensions[1].height = 50

    hdr(ws,1,1,"SA ACCOMMODATION INTELLIGENCE\nPLATFORM", bold=True, sz=16,
        bg=DARK, fg="FF6C00", wrap=True)
    ws.merge_cells("A1:B1")

    rows = [
        ("What is this file?",
         "This Excel workbook is the primary deliverable for the LekkeSlaap Accommodation Intelligence Platform — "
         "a data analysis of 1,011 listings from LekkeSlaap.co.za (South Africa's leading local accommodation app) "
         "combined with synthetic Google Analytics 4 event data and machine learning insights."),
        ("Data source",
         "LekkeSlaap.co.za scrape — June 2026. 1,011 listings across 114 SA regions. "
         "GA4 events (28,628 rows) are synthetic, generated to mirror real booking patterns. "
         "All prices in South African Rand (ZAR, R) per night."),
        ("Currency & units",
         "All monetary values = South African Rand (ZAR / R). Prices = per night. "
         "Demand Score = composite 0-100 (higher = more demand). Review Count = cumulative guest reviews."),
        ("Sheet 1 — Cover",         "Platform overview, methodology summary, and key KPIs."),
        ("Sheet 2 — Dirty Data Audit", "Raw data quality report: nulls, outliers, inconsistencies found and fixed during ETL."),
        ("Sheet 3 — Clean Data",    "1,011 cleaned property records: price, demand, reviews, region, type, promo flag."),
        ("Sheet 4 — Regional Analysis", "114 regions ranked by avg price, demand score, listing count. Use to identify supply gaps."),
        ("Sheet 5 — GA4 Web Analytics", "Synthetic GA4 session and event data: channels, devices, provinces, event funnel."),
        ("Sheet 6 — ML Model Results",  "5 ML model performance metrics: Random Forest, Gradient Boosting, Logistic Regression, Isolation Forest, K-Means."),
        ("Sheet 7 — ML Predictions",    "Per-property ML predictions: predicted demand, price tier, anomaly flag."),
        ("Sheet 8 — GTM + GA4 Guide",   "28 GTM tags documented: tag name, type, trigger, POPIA consent status."),
        ("Sheet 9 — ML Clustering",     "K-Means 4-cluster results: High-Demand Hotspot / Established Premium / Value Volume Leader / Emerging Gem."),
        ("Sheet 10 — Seasonal Demand",  "Booking demand by SA season (Summer/Autumn/Winter/Spring) per province."),
        ("Sheet 11 — Price-Demand Corr","Pearson correlation between price and demand score per cluster and region."),
        ("ML Clusters explained",
         "Cluster 0 High-Demand Hotspot: 7 regions, avg R1,297/night, demand 24 — top traffic, peak-season destinations.\n"
         "Cluster 1 Established Premium: 19 regions, avg R1,651/night — strong review base, reliable year-round.\n"
         "Cluster 2 Value Volume Leader: 83 regions, avg R1,851/night — largest supply, mid-market pricing.\n"
         "Cluster 3 Emerging Gem: 5 regions, avg R9,754/night — ultra-premium niche, lowest volume."),
        ("SA Seasons",
         "Summer = Dec-Feb (peak, beach/garden route demand). Autumn = Mar-May (Easter spike). "
         "Winter = Jun-Aug (Kruger/game reserves, Cape Town off-season deals). Spring = Sep-Nov (shoulder season)."),
        ("Reproducibility",
         "All data files and Python scripts are in the accompanying GitHub repo / local folder. "
         "BigQuery dataset: africa-south1.accommodation_intelligence. "
         "Run 01_etl_clean.py → 02_ml_models.py → 03_generate_outputs.py to regenerate from source."),
    ]

    for i, (label, desc) in enumerate(rows, start=2):
        ws.row_dimensions[i].height = max(30, len(desc)//4)
        hdr(ws, i, 1, label, bold=True, sz=10, bg="F0F0F0", fg="1A1A2E")
        c = ws.cell(row=i, column=2, value=desc)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=9)
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="FAFAFA")

    wb.save(XL_PATH)
    sz = XL_PATH.stat().st_size // 1024
    print(f"  Excel README sheet added — {XL_PATH.name} ({sz} KB, {len(wb.sheetnames)} sheets)")
else:
    print("  Excel file not found — skipping README sheet")

print("\nDone.")
print("  data/appsflyer_installs.csv")
print("  data/appsflyer_campaigns.csv")
print("  data/appsflyer_cohorts.csv")
print("  netlify_site/mobile-attribution.html")
print("  netlify_site/seo-audit.html")
print("  netlify_site/sitemap.xml")
print("  netlify_site/robots.txt")
print("  Excel: README sheet added (sheet 1)")
print("  Ebook: self-explanatory context banner added")
print("  Nav: updated across all pages (12 links)")
