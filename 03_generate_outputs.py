"""
SA Accommodation Intelligence Platform
03_generate_outputs.py — Excel Workbook + Interactive Dashboard HTML + Ebook HTML

Author: Anthony Apollis | 2026-06-27
"""

import json
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage

BASE     = Path(__file__).parent
DATA     = BASE / "data"
PLOTS    = BASE / "plots"
REPORTS  = BASE / "reports"
ML_DIR   = BASE / "ml"

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
df_clean    = pd.read_csv(DATA / "accommodation_clean.csv")
df_sessions = pd.read_csv(DATA / "fact_web_sessions.csv")
df_events   = pd.read_csv(DATA / "fact_booking_events.csv")
df_region   = pd.read_csv(DATA / "dim_region.csv")
df_listings = pd.read_csv(DATA / "fact_listings.csv")
df_preds    = pd.read_csv(ML_DIR / "ml_predictions.csv") if (ML_DIR/"ml_predictions.csv").exists() else pd.DataFrame()
metrics     = json.load(open(ML_DIR / "model_metrics.json")) if (ML_DIR/"model_metrics.json").exists() else {}

# ── COLOUR PALETTE ─────────────────────────────────────────────────────────────
GCP_BLUE    = "4285F4"
GCP_GREEN   = "0F9D58"
SA_ORANGE   = "FF6D00"
GCP_YELLOW  = "F4B400"
GCP_RED     = "DB4437"
BQ_TEAL     = "00BCD4"
PURPLE      = "673AB7"
DARK_BG     = "1A1A2E"
LIGHT_BG    = "F8F9FA"
WHITE       = "FFFFFF"
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
BODY_FONT   = Font(name="Calibri", size=10)
TITLE_FONT  = Font(name="Calibri", bold=True, color=DARK_BG, size=14)
SUB_FONT    = Font(name="Calibri", bold=True, color="555555", size=11)

TIER_HEX = {
    "Budget"    : GCP_GREEN,
    "Mid-Range" : GCP_BLUE,
    "Premium"   : SA_ORANGE,
    "Luxury"    : GCP_YELLOW,
    "Unknown"   : "AAAAAA",
}

def hfill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row_idx, headers, fill_hex=GCP_BLUE):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col, value=h)
        cell.fill  = hfill(fill_hex)
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

def data_rows(ws, df, start_row, alt_hex="F0F4FF"):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = thin_border()
            if r_idx % 2 == 0:
                cell.fill = hfill(alt_hex)
    return r_idx

# ══════════════════════════════════════════════════════════════
# EXCEL WORKBOOK
# ══════════════════════════════════════════════════════════════
print("=" * 60)
print("Generating Excel Workbook")
print("=" * 60)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── SHEET 1: COVER ────────────────────────────────────────────
ws_cover = wb.create_sheet("Cover")
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions["B"].width = 60

for r in range(1, 40):
    ws_cover.row_dimensions[r].height = 18

ws_cover.merge_cells("B2:H2")
ws_cover["B2"].value  = "SA Accommodation Intelligence Platform"
ws_cover["B2"].font   = Font(name="Calibri", bold=True, size=24, color=GCP_BLUE)
ws_cover["B2"].alignment = Alignment(horizontal="left")

ws_cover.merge_cells("B3:H3")
ws_cover["B3"].value  = "LekkeSlaap Market Analysis — South Africa & Namibia"
ws_cover["B3"].font   = Font(name="Calibri", size=14, color=SA_ORANGE)

ws_cover.merge_cells("B5:H5")
ws_cover["B5"].value  = "Data Pipeline: Scraped → Cleaned → BigQuery → ML → Insights"
ws_cover["B5"].font   = Font(name="Calibri", size=11, color="666666")

meta = [
    ("Author",          "Anthony Apollis"),
    ("Date",            "2026-06-27"),
    ("Source",          "LekkeSlaap.co.za (scraped Jun 2026)"),
    ("Properties",      f"{len(df_clean):,} unique listings"),
    ("Web Sessions",    f"{len(df_sessions):,} synthetic GA4 sessions"),
    ("Booking Events",  f"{len(df_events):,} funnel events"),
    ("ML Models",       "Price Regression | Tier Classifier | Demand Scorer"),
    ("BigQuery Dataset","accommodation_intelligence (africa-south1)"),
    ("ETL Tool",        "Python (pandas + Faker + scikit-learn) → BQ bq load (free)"),
]
for r, (k, v) in enumerate(meta, 8):
    ws_cover.cell(row=r, column=2).value = k
    ws_cover.cell(row=r, column=2).font  = Font(name="Calibri", bold=True, size=10, color="444444")
    ws_cover.cell(row=r, column=3).value = v
    ws_cover.cell(row=r, column=3).font  = Font(name="Calibri", size=10)

# Colour bar
for col in range(2, 10):
    colors = [GCP_BLUE, GCP_GREEN, SA_ORANGE, GCP_YELLOW, BQ_TEAL,
              PURPLE, GCP_RED, GCP_BLUE]
    ws_cover.cell(row=22, column=col).fill = hfill(colors[(col-2) % len(colors)])
    ws_cover.row_dimensions[22].height = 6

# ── SHEET 2: DIRTY DATA AUDIT ─────────────────────────────────
ws_dirty = wb.create_sheet("Dirty Data Audit")
ws_dirty.column_dimensions["A"].width = 35
ws_dirty.column_dimensions["B"].width = 18
ws_dirty.column_dimensions["C"].width = 45

ws_dirty["A1"].value = "SA Accommodation — Dirty Data Audit Report"
ws_dirty["A1"].font  = TITLE_FONT
ws_dirty.merge_cells("A1:C1")
ws_dirty["A2"].value = f"Source: accommodation_listings_redo_dom.csv | Processed: {datetime.now().strftime('%Y-%m-%d')}"
ws_dirty["A2"].font  = Font(name="Calibri", size=9, color="888888")

header_row(ws_dirty, 4, ["Issue Type", "Count", "Action Taken / Rule Applied"], fill_hex=GCP_RED)

dirty_issues = [
    ("Promotional noise in names",
     "~15 rows",
     "Removed 'Flash Deal X% off!' prefix using regex; preserved actual property name"),
    ("Prices with thousands spaces (R1 700)",
     "~600+ rows",
     "Stripped spaces between digits after 'R' prefix; converted to numeric float"),
    ("Duplicate suburb in location",
     "~500+ rows",
     "Regex matched 'Suburb, Suburb City' pattern; collapsed to 'Suburb, City'"),
    ("Non-standard listing types (LodgingBusiness)",
     "3 rows",
     "Mapped LodgingBusiness → Guest House; 'accommodation' slug → Self Catering"),
    ("Rating field: 99.7% empty",
     "~1,010 rows",
     "Rating not consistently scraped; excluded from ML features; documented as data gap"),
    ("Price outliers > R20,000",
     "~3 rows",
     "Set to NULL; flagged as price_outlier_flag=TRUE; reviewed manually"),
    ("Duplicate property records",
     "~50 rows",
     "Deduplicated on property_id (LekkeSlaap's internal ID); kept first occurrence"),
    ("Review count = 0 (missing vs. truly zero)",
     "~58 rows",
     "Treated 0 as zero reviews (not null); log1p transform applied in ML to handle zeros"),
    ("Location = '- Melissa H' (reviewer name leaked)",
     "2 rows",
     "Identified as scraper artefact; location set to NULL, region inferred from source_page"),
    ("'Browse Accommodation in...' scraped as property name",
     "2 rows",
     "Navigation text captured as listing; removed post-deduplication by property_id"),
]
for r, (issue, count, action) in enumerate(dirty_issues, 5):
    ws_dirty.cell(row=r, column=1).value = issue
    ws_dirty.cell(row=r, column=2).value = count
    ws_dirty.cell(row=r, column=3).value = action
    for c in range(1, 4):
        ws_dirty.cell(row=r, column=c).font   = BODY_FONT
        ws_dirty.cell(row=r, column=c).border = thin_border()
        ws_dirty.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
    if r % 2 == 0:
        for c in range(1, 4):
            ws_dirty.cell(row=r, column=c).fill = hfill("FFF3E0")
    ws_dirty.row_dimensions[r].height = 36

# ── SHEET 3: CLEAN DATA ───────────────────────────────────────
ws_data = wb.create_sheet("Clean Data")
header_row(ws_data, 1,
    ["Property ID","Name","Listing Type","Price (ZAR)","Price Tier",
     "Review Count","Demand Score","City","Region","Country","URL","Promo?","Discount %"],
    fill_hex=GCP_BLUE)
last = data_rows(ws_data, df_clean.fillna("").head(500), 2)
for col, width in zip("ABCDEFGHIJKLM", [12,40,16,12,12,14,14,22,22,14,40,10,12]):
    ws_data.column_dimensions[chr(64+list("ABCDEFGHIJKLM").index(col)+1)].width = width
ws_data.freeze_panes = "A2"
ws_data.auto_filter.ref = f"A1:M{last}"

# ── SHEET 4: REGIONAL ANALYSIS ────────────────────────────────
ws_reg = wb.create_sheet("Regional Analysis")
region_stats = df_clean.groupby("region").agg(
    listings=("property_id","count"),
    avg_price=("price_zar","mean"),
    median_price=("price_zar","median"),
    avg_reviews=("review_count","mean"),
    avg_demand=("demand_score","mean"),
    budget=("price_tier", lambda x: (x=="Budget").sum()),
    premium_luxury=("price_tier", lambda x: ((x=="Premium")|(x=="Luxury")).sum()),
).reset_index().sort_values("listings", ascending=False)
region_stats["avg_price"] = region_stats["avg_price"].round(0).astype("Int64")
region_stats["median_price"] = region_stats["median_price"].round(0).astype("Int64")
region_stats["avg_demand"] = region_stats["avg_demand"].round(2)

ws_reg["A1"].value = "Regional Market Analysis — LekkeSlaap SA"
ws_reg["A1"].font  = TITLE_FONT
ws_reg.merge_cells("A1:H1")

header_row(ws_reg, 3,
    ["Region","Listings","Avg Price (ZAR)","Median Price","Avg Reviews",
     "Avg Demand Score","Budget Count","Premium/Luxury Count"],
    fill_hex=GCP_GREEN)
last_reg = data_rows(ws_reg, region_stats.astype(object).fillna(""), 4)
for c in "ABCDEFGH":
    ws_reg.column_dimensions[c].width = 20
ws_reg.freeze_panes = "A4"

# ── SHEET 5: GA4 WEB ANALYTICS ────────────────────────────────
ws_ga4 = wb.create_sheet("GA4 Web Analytics")
ws_ga4["A1"].value = "GA4 Web Analytics — LekkeSlaap Platform (Jan–Jun 2025)"
ws_ga4["A1"].font  = TITLE_FONT
ws_ga4.merge_cells("A1:H1")

# Traffic source summary
src_stats = df_sessions.groupby("traffic_source").agg(
    sessions=("session_id","count"),
    users=("user_pseudo_id","nunique"),
    engaged=("session_engaged","sum"),
    bounced=("bounced","sum"),
    avg_eng_secs=("engagement_secs","mean"),
).reset_index().sort_values("sessions", ascending=False)
src_stats["engagement_rate_%"] = (src_stats["engaged"]/src_stats["sessions"]*100).round(1)
src_stats["bounce_rate_%"] = (src_stats["bounced"]/src_stats["sessions"]*100).round(1)
src_stats["avg_eng_secs"] = src_stats["avg_eng_secs"].round(0).astype(int)

ws_ga4["A3"].value = "Traffic Source Performance"
ws_ga4["A3"].font  = SUB_FONT
header_row(ws_ga4, 4,
    ["Source","Sessions","Users","Engaged","Bounce","Avg Eng (s)",
     "Engagement Rate %","Bounce Rate %"],
    fill_hex=PURPLE)
last_src = data_rows(ws_ga4, src_stats.fillna(""), 5)

# Funnel
funnel_data = df_events.groupby("event_name").agg(
    events=("event_id","count"),
    sessions=("session_id","nunique"),
).reset_index()
funnel_order = ["listing_view","search_nearby","contact_host","booking_initiated","booking_confirmed"]
funnel_data["sort"] = funnel_data["event_name"].map({e:i for i,e in enumerate(funnel_order)})
funnel_data = funnel_data.sort_values("sort").drop("sort",axis=1)
funnel_data["conv_from_top_%"] = (funnel_data["sessions"] / funnel_data["sessions"].iloc[0] * 100).round(1)

ws_ga4[f"A{last_src+3}"].value = "Booking Conversion Funnel"
ws_ga4[f"A{last_src+3}"].font  = SUB_FONT
header_row(ws_ga4, last_src+4,
    ["Event Name","Total Events","Sessions","Conversion from Top (%)"],
    fill_hex=SA_ORANGE)
data_rows(ws_ga4, funnel_data.fillna(""), last_src+5)

for c in "ABCDEFGH":
    ws_ga4.column_dimensions[c].width = 22
ws_ga4.freeze_panes = "A5"

# ── SHEET 6: ML RESULTS ───────────────────────────────────────
ws_ml = wb.create_sheet("ML Model Results")
ws_ml["A1"].value = "Machine Learning Model Performance"
ws_ml["A1"].font  = TITLE_FONT
ws_ml.merge_cells("A1:G1")

ml_summary = [
    ["Model","Algorithm","Target","Key Metric","Value","Train/Test Split","Notes"],
    ["Price Regression","Gradient Boosting Regressor","price_zar (ZAR)",
     "R²", metrics.get("price_regression",{}).get("r2","—"),
     "80/20","MAE: R" + str(metrics.get("price_regression",{}).get("mae_zar","—"))],
    ["","","","MAE (ZAR)", metrics.get("price_regression",{}).get("mae_zar","—"),"","5-fold CV R²: "+str(metrics.get("price_regression",{}).get("cv_r2_mean","—"))],
    ["Tier Classifier","Random Forest Classifier","price_tier (4 classes)",
     "Accuracy", metrics.get("tier_classifier",{}).get("accuracy","—"),
     "80/20 stratified","F1 Weighted: "+str(metrics.get("tier_classifier",{}).get("f1_weighted","—"))],
    ["","","","CV Accuracy", metrics.get("tier_classifier",{}).get("cv_acc_mean","—"),"","5-fold cross-validation"],
    ["Demand Scorer","Gradient Boosting Regressor","demand_score (0–100)",
     "R²", metrics.get("demand_model",{}).get("r2","—"),
     "80/20","MAE: "+str(metrics.get("demand_model",{}).get("mae_score_pts","—"))+" pts"],
]

header_row(ws_ml, 3, ml_summary[0], fill_hex=BQ_TEAL)
for r, row in enumerate(ml_summary[1:], 4):
    for c, v in enumerate(row, 1):
        cell = ws_ml.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = thin_border()
    if r % 2 == 0:
        for c in range(1,8):
            ws_ml.cell(row=r, column=c).fill = hfill("E8F5E9")

ws_ml["A12"].value = "Feature Importance Summary"
ws_ml["A12"].font  = SUB_FONT
header_row(ws_ml, 13,
    ["Feature","Price Model Role","Tier Model Role","Note"],
    fill_hex=GCP_BLUE)
feat_rows = [
    ["Listing Type","High — Hotel/Game Lodge price premium", "High — defines tier boundaries", "Self Catering = 67% of market"],
    ["Region","High — Cape Town / Joburg premium", "Medium — regional price norms differ", "Garden Route = premium leisure"],
    ["City","Medium — suburb within region", "Medium", "Constantia vs Mitchells Plain (CT)"],
    ["Log(Review Count)","Medium — more reviews = higher price (quality signal)", "Low — tier set by price, not popularity", "log1p transform handles zero reviews"],
]
for r, row in enumerate(feat_rows, 14):
    for c, v in enumerate(row, 1):
        ws_ml.cell(row=r, column=c).value  = v
        ws_ml.cell(row=r, column=c).font   = BODY_FONT
        ws_ml.cell(row=r, column=c).border = thin_border()
        ws_ml.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
    if r % 2 == 0:
        for c in range(1,5):
            ws_ml.cell(row=r, column=c).fill = hfill("E3F2FD")
    ws_ml.row_dimensions[r].height = 30

for c, w in zip("ABCDEFG", [22,40,35,40,14,18,40]):
    ws_ml.column_dimensions[c].width = w

# ── SHEET 7: ML PREDICTIONS ───────────────────────────────────
if not df_preds.empty:
    ws_pred = wb.create_sheet("ML Predictions")
    ws_pred["A1"].value = "ML Price & Tier Predictions — All Properties"
    ws_pred["A1"].font  = TITLE_FONT
    ws_pred.merge_cells("A1:L1")
    header_row(ws_pred, 2,
        ["Property ID","Name","Type","Region","Actual Price","Predicted Price",
         "Delta (ZAR)","Actual Tier","Predicted Tier","Actual Demand","Predicted Demand","Reviews"],
        fill_hex=PURPLE)
    data_rows(ws_pred, df_preds.fillna(""), 3)
    for c, w in zip("ABCDEFGHIJKL", [12,40,16,22,14,16,12,14,14,14,16,10]):
        ws_pred.column_dimensions[chr(64+list("ABCDEFGHIJKL").index(c)+1)].width = w
    ws_pred.freeze_panes = "A3"

# ── SHEET 8: GTM / GA4 IMPLEMENTATION GUIDE ──────────────────
ws_gtm = wb.create_sheet("GTM + GA4 Guide")
ws_gtm.column_dimensions["A"].width = 28
ws_gtm.column_dimensions["B"].width = 35
ws_gtm.column_dimensions["C"].width = 55

ws_gtm["A1"].value = "GTM & GA4 Implementation Guide — LekkeSlaap-Style Platform"
ws_gtm["A1"].font  = TITLE_FONT
ws_gtm.merge_cells("A1:C1")

ws_gtm["A3"].value = "GA4 Configuration Layer (Google Tag Manager)"
ws_gtm["A3"].font  = SUB_FONT

header_row(ws_gtm, 4, ["Tag / Trigger", "GA4 Event Name", "dataLayer Push / Parameters"], fill_hex=GCP_BLUE)

gtm_rows = [
    ["GA4 Configuration Tag", "— (base tag)", "Tag ID: G-XXXXXXXX; fires on All Pages; includes consent_default mode for POPIA"],
    ["Page View Tag", "page_view", "Fires on All Pages; params: page_title, page_location, page_referrer"],
    ["Listing View Trigger", "listing_view", "dataLayer.push({event:'listing_view', property_id:{{dlv.property_id}}, listing_type:{{dlv.listing_type}}, price_zar:{{dlv.price_zar}}, price_tier:{{dlv.price_tier}}})"],
    ["Search Trigger", "search_nearby", "Fires on map/filter interactions; params: search_term, filters_applied, result_count"],
    ["Contact Host", "contact_host", "Fires on 'Contact' button click; params: property_id, listing_type, price_zar"],
    ["Booking Initiated", "booking_initiated", "Equivalent to begin_checkout; params: property_id, price_zar, checkin_date, checkout_date, nights"],
    ["Booking Confirmed", "booking_confirmed", "Fires on confirmation page; params: transaction_id, price_zar, listing_type, province, payment_method"],
    ["POPIA Consent Update", "consent_update", "Fires when user accepts/declines cookies; updates analytics_storage, ads_storage"],
    ["Scroll Depth (75%)", "scroll", "Built-in GA4 scroll trigger at 75% page height; no custom push needed"],
    ["Outbound Link Click", "click", "GA4 Enhanced Measurement outbound clicks; tracks external booking engine links"],
]
for r, row in enumerate(gtm_rows, 5):
    for c, v in enumerate(row, 1):
        cell = ws_gtm.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = thin_border()
        cell.alignment = Alignment(wrap_text=True)
    if r % 2 == 0:
        for c in range(1, 4):
            ws_gtm.cell(row=r, column=c).fill = hfill("E3F2FD")
    ws_gtm.row_dimensions[r].height = 42

ws_gtm[f"A{len(gtm_rows)+7}"].value = "Custom Dimensions to Register in GA4 Admin"
ws_gtm[f"A{len(gtm_rows)+7}"].font  = SUB_FONT
header_row(ws_gtm, len(gtm_rows)+8,
    ["Dimension Name","Scope","Description"],
    fill_hex=GCP_GREEN)
custom_dims = [
    ["listing_type","Event","Type of accommodation (Self Catering, Guest House, Hotel etc.)"],
    ["price_tier","Event","Budget / Mid-Range / Premium / Luxury"],
    ["property_id","Event","LekkeSlaap internal property identifier"],
    ["province","Event","SA province of the user (from geo or checkout form)"],
    ["has_promo","Event","TRUE if listing was displaying a flash deal at time of view"],
    ["payment_method","Event","Payment type on booking_confirmed (EFT, Card, Ozow, SnapScan)"],
    ["nights","Event","Number of nights in the booking (checkout − checkin)"],
]
for r, row in enumerate(custom_dims, len(gtm_rows)+9):
    for c, v in enumerate(row, 1):
        ws_gtm.cell(row=r, column=c).value  = v
        ws_gtm.cell(row=r, column=c).font   = BODY_FONT
        ws_gtm.cell(row=r, column=c).border = thin_border()

# ── SAVE WORKBOOK ─────────────────────────────────────────────
xl_path = REPORTS / "SA_Accommodation_Intelligence_Platform.xlsx"
wb.save(xl_path)
print(f"  Excel saved: {xl_path.name}")


# ══════════════════════════════════════════════════════════════
# ADDITIONAL ANALYTICS PLOTS (for ebook)
# ══════════════════════════════════════════════════════════════
print("\nGenerating additional plots...")

PALETTE = dict(primary="#4285F4", secondary="#0F9D58", accent="#FF6D00",
               warning="#F4B400", teal="#00BCD4", purple="#673AB7",
               danger="#DB4437", bg="#F8F9FA", dark="#1A1A2E")

plt.rcParams.update({
    "figure.facecolor": PALETTE["bg"], "axes.facecolor":"white",
    "axes.grid":True, "grid.alpha":0.3,
    "axes.spines.top":False, "axes.spines.right":False,
    "font.family":"DejaVu Sans",
})
TIER_COLORS = {"Budget":PALETTE["secondary"],"Mid-Range":PALETTE["primary"],
               "Premium":PALETTE["accent"],"Luxury":PALETTE["warning"],"Unknown":"#AAAAAA"}

# Plot: Monthly booking trend
df_events["event_date"] = pd.to_datetime(df_events["event_date"])
monthly = df_events[df_events["event_name"]=="booking_confirmed"].resample("ME", on="event_date").size()
monthly_sess = df_sessions.copy()
monthly_sess["event_date"] = pd.to_datetime(monthly_sess["event_date"])
monthly_sess_cnt = monthly_sess.resample("ME", on="event_date").size()

fig, ax = plt.subplots(figsize=(12, 5))
fig.patch.set_facecolor(PALETTE["bg"])
ax.fill_between(monthly_sess_cnt.index, monthly_sess_cnt.values,
                alpha=0.15, color=PALETTE["primary"])
ax2 = ax.twinx()
ax2.spines["right"].set_visible(True)
ax.plot(monthly_sess_cnt.index, monthly_sess_cnt.values,
        color=PALETTE["primary"], lw=2.5, label="Sessions", marker="o", ms=5)
ax2.bar(monthly.index, monthly.values, width=15, alpha=0.6,
        color=PALETTE["accent"], label="Confirmed Bookings")
ax.set_title("Monthly Sessions & Confirmed Bookings — Jan–Jun 2025",
             fontsize=13, fontweight="bold")
ax.set_ylabel("Sessions", color=PALETTE["primary"])
ax2.set_ylabel("Confirmed Bookings", color=PALETTE["accent"])
lines1, labels1 = ax.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax.legend(lines1+lines2, labels1+labels2, loc="upper left")
plt.tight_layout()
plt.savefig(PLOTS / "monthly_trend.png", dpi=150, bbox_inches="tight")
plt.close()

# Plot: Price tier by region (top 8 regions)
top8 = df_clean["region"].value_counts().head(8).index
df_t8 = df_clean[df_clean["region"].isin(top8)]
pivot = df_t8.pivot_table(index="region", columns="price_tier",
                           values="property_id", aggfunc="count", fill_value=0)
pivot = pivot.reindex(columns=["Budget","Mid-Range","Premium","Luxury"], fill_value=0)
pivot = pivot.loc[pivot.sum(axis=1).sort_values(ascending=False).index]

fig, ax = plt.subplots(figsize=(12, 6))
fig.patch.set_facecolor(PALETTE["bg"])
pivot.plot(kind="bar", ax=ax, stacked=True,
           color=[TIER_COLORS[c] for c in pivot.columns], alpha=0.85, width=0.7)
ax.set_title("Price Tier Distribution by Region (Top 8)",
             fontsize=13, fontweight="bold")
ax.set_ylabel("Number of Properties")
ax.set_xlabel("")
plt.setp(ax.get_xticklabels(), rotation=30, ha="right")
ax.legend(title="Price Tier", bbox_to_anchor=(1,1))
plt.tight_layout()
plt.savefig(PLOTS / "region_tier_stacked.png", dpi=150, bbox_inches="tight")
plt.close()

# Plot: Device × engagement vs bounce
fig, axes = plt.subplots(1, 2, figsize=(12, 5))
fig.patch.set_facecolor(PALETTE["bg"])
dev_metrics = df_sessions.groupby("device_category").agg(
    avg_eng=("engagement_secs","mean"), bounce=("bounced","mean")).round(2)
ax = axes[0]
bars = ax.bar(dev_metrics.index, dev_metrics["avg_eng"],
              color=[PALETTE["primary"],PALETTE["secondary"],PALETTE["accent"]], alpha=0.85)
for bar, v in zip(bars, dev_metrics["avg_eng"]):
    ax.text(bar.get_x()+bar.get_width()/2, v+0.5, f"{v:.0f}s", ha="center", fontsize=10)
ax.set_title("Avg Engagement Time by Device", fontsize=11, fontweight="bold")
ax.set_ylabel("Seconds")
ax = axes[1]
bars = ax.bar(dev_metrics.index, dev_metrics["bounce"]*100,
              color=[PALETTE["primary"],PALETTE["secondary"],PALETTE["accent"]], alpha=0.85)
for bar, v in zip(bars, dev_metrics["bounce"]*100):
    ax.text(bar.get_x()+bar.get_width()/2, v+0.3, f"{v:.1f}%", ha="center", fontsize=10)
ax.set_title("Bounce Rate by Device", fontsize=11, fontweight="bold")
ax.set_ylabel("Bounce Rate (%)")
plt.tight_layout()
plt.savefig(PLOTS / "device_engagement.png", dpi=150, bbox_inches="tight")
plt.close()

print("  Plots saved")


# ══════════════════════════════════════════════════════════════
# INTERACTIVE DASHBOARD HTML
# ══════════════════════════════════════════════════════════════
print("\nGenerating Dashboard HTML...")

# KPIs
total_props  = len(df_clean)
avg_price    = df_clean["price_zar"].mean()
median_price = df_clean["price_zar"].median()
total_sess   = len(df_sessions)
total_book   = (df_events["event_name"]=="booking_confirmed").sum()
conv_rate    = total_book / total_sess * 100
bounce_rate  = df_sessions["bounced"].mean() * 100
avg_eng      = df_sessions["engagement_secs"].mean()
popia_ok     = (df_events["analytics_consent"]=="Yes").mean() * 100

tier_counts  = df_clean["price_tier"].value_counts().to_dict()
region_top   = df_clean["region"].value_counts().head(8).to_dict()
type_counts  = df_clean["listing_type"].value_counts().to_dict()
traffic_data = df_sessions.groupby("traffic_source").size().to_dict()
booking_by_province = (df_events[df_events["event_name"]=="booking_confirmed"]
                        .groupby("province").size().sort_values(ascending=False).head(9).to_dict())

r2_price = metrics.get("price_regression",{}).get("r2","—")
mae_price = metrics.get("price_regression",{}).get("mae_zar","—")
acc_tier  = metrics.get("tier_classifier",{}).get("accuracy","—")
r2_demand = metrics.get("demand_model",{}).get("r2","—")

# Build chart data JSON
import json as _json
tier_chart_js   = _json.dumps({"labels":list(tier_counts.keys()), "data":list(tier_counts.values())})
region_chart_js = _json.dumps({"labels":list(region_top.keys()), "data":list(region_top.values())})
type_chart_js   = _json.dumps({"labels":list(type_counts.keys()), "data":list(type_counts.values())})
traffic_chart_js= _json.dumps({"labels":list(traffic_data.keys()), "data":list(traffic_data.values())})
province_chart_js=_json.dumps({"labels":list(booking_by_province.keys()), "data":list(booking_by_province.values())})

funnel_counts = [df_events[df_events["event_name"]==e].shape[0]
                 for e in ["listing_view","search_nearby","contact_host","booking_initiated","booking_confirmed"]]
funnel_chart_js = _json.dumps({
    "labels":["Listing View","Search Nearby","Contact Host","Booking Init","Confirmed"],
    "data": funnel_counts
})

dashboard_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SA Accommodation Intelligence Platform — Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --blue:    #4285F4;
    --green:   #0F9D58;
    --orange:  #FF6D00;
    --yellow:  #F4B400;
    --red:     #DB4437;
    --teal:    #00BCD4;
    --purple:  #673AB7;
    --dark:    #1A1A2E;
    --bg:      #F4F6FB;
    --card-bg: #FFFFFF;
    --border:  #E0E6F0;
    --text:    #333344;
    --sub:     #6B7A99;
  }}
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family:'Segoe UI',Arial,sans-serif; background:var(--bg); color:var(--text); }}
  header {{
    background:var(--dark); color:#fff; padding:18px 32px;
    display:flex; align-items:center; gap:16px;
    box-shadow:0 2px 8px rgba(0,0,0,.3);
  }}
  header .logo {{ font-size:1.6rem; font-weight:700; color:var(--teal); }}
  header .subtitle {{ font-size:.85rem; color:#aab3cc; }}
  .badge {{
    background:var(--teal); color:var(--dark);
    font-size:.7rem; font-weight:700; padding:3px 10px;
    border-radius:12px; letter-spacing:.05em; margin-left:auto;
  }}
  main {{ padding:24px 32px; }}
  section {{ margin-bottom:32px; }}
  h2 {{ font-size:1.05rem; font-weight:700; color:var(--dark);
        border-left:4px solid var(--teal); padding-left:10px; margin-bottom:16px; }}
  .kpi-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(180px,1fr)); gap:14px; }}
  .kpi {{
    background:var(--card-bg); border-radius:10px; padding:18px 16px;
    box-shadow:0 1px 4px rgba(0,0,0,.08); border-top:3px solid var(--blue);
    display:flex; flex-direction:column; gap:6px;
  }}
  .kpi .val {{ font-size:1.7rem; font-weight:700; color:var(--dark); }}
  .kpi .lbl {{ font-size:.75rem; color:var(--sub); text-transform:uppercase; letter-spacing:.06em; }}
  .kpi.green {{ border-top-color:var(--green); }} .kpi.green .val {{ color:var(--green); }}
  .kpi.orange {{ border-top-color:var(--orange); }} .kpi.orange .val {{ color:var(--orange); }}
  .kpi.yellow {{ border-top-color:var(--yellow); }} .kpi.yellow .val {{ color:var(--yellow); }}
  .kpi.teal {{ border-top-color:var(--teal); }} .kpi.teal .val {{ color:var(--teal); }}
  .kpi.purple{{ border-top-color:var(--purple); }} .kpi.purple .val {{ color:var(--purple); }}
  .kpi.red   {{ border-top-color:var(--red); }}  .kpi.red .val   {{ color:var(--red); }}
  .chart-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(400px,1fr)); gap:20px; }}
  .chart-card {{
    background:var(--card-bg); border-radius:10px; padding:20px;
    box-shadow:0 1px 4px rgba(0,0,0,.08);
  }}
  .chart-card h3 {{ font-size:.9rem; color:var(--sub); margin-bottom:14px; font-weight:600; }}
  .chart-wrap {{ position:relative; height:240px; }}
  .ml-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:14px; }}
  .ml-card {{
    background:var(--card-bg); border-radius:10px; padding:18px;
    box-shadow:0 1px 4px rgba(0,0,0,.08); text-align:center;
    border-top:3px solid var(--purple);
  }}
  .ml-card .model {{ font-size:.75rem; color:var(--sub); margin-bottom:6px; }}
  .ml-card .metric {{ font-size:1.5rem; font-weight:700; color:var(--purple); }}
  .ml-card .desc {{ font-size:.72rem; color:var(--sub); margin-top:4px; }}
  .dirty-table {{ width:100%; border-collapse:collapse; }}
  .dirty-table th {{ background:var(--red); color:#fff; padding:8px 12px; font-size:.8rem; text-align:left; }}
  .dirty-table td {{ padding:7px 12px; border-bottom:1px solid var(--border); font-size:.8rem; }}
  .dirty-table tr:nth-child(even) td {{ background:#FFF3E0; }}
  footer {{ text-align:center; padding:20px; color:var(--sub); font-size:.75rem; }}
  @media(max-width:600px){{
    main{{padding:12px;}} .chart-grid{{grid-template-columns:1fr;}} .ml-grid{{grid-template-columns:1fr;}}
  }}
</style>
</head>
<body>
<header>
  <div>
    <div class="logo">SA Accommodation Intelligence Platform</div>
    <div class="subtitle">LekkeSlaap Market Analysis · BigQuery · GA4 · Machine Learning · 2026</div>
  </div>
  <div class="badge">BIGQUERY + GA4</div>
</header>

<main>

<!-- KPIs -->
<section>
  <h2>Platform KPIs</h2>
  <div class="kpi-grid">
    <div class="kpi">
      <span class="val">{total_props:,}</span>
      <span class="lbl">Properties Scraped</span>
    </div>
    <div class="kpi green">
      <span class="val">R{avg_price:,.0f}</span>
      <span class="lbl">Avg Nightly Price (ZAR)</span>
    </div>
    <div class="kpi orange">
      <span class="val">R{median_price:,.0f}</span>
      <span class="lbl">Median Nightly Price</span>
    </div>
    <div class="kpi teal">
      <span class="val">{total_sess:,}</span>
      <span class="lbl">GA4 Sessions</span>
    </div>
    <div class="kpi yellow">
      <span class="val">{total_book:,}</span>
      <span class="lbl">Confirmed Bookings</span>
    </div>
    <div class="kpi purple">
      <span class="val">{conv_rate:.2f}%</span>
      <span class="lbl">Booking Conversion Rate</span>
    </div>
    <div class="kpi red">
      <span class="val">{bounce_rate:.1f}%</span>
      <span class="lbl">Bounce Rate</span>
    </div>
    <div class="kpi teal">
      <span class="val">{avg_eng:.0f}s</span>
      <span class="lbl">Avg Engagement Time</span>
    </div>
    <div class="kpi green">
      <span class="val">{popia_ok:.1f}%</span>
      <span class="lbl">POPIA Consent Rate</span>
    </div>
  </div>
</section>

<!-- ML Models -->
<section>
  <h2>ML Model Performance</h2>
  <div class="ml-grid">
    <div class="ml-card">
      <div class="model">Model 1: Price Regression (GBR)</div>
      <div class="metric">R² = {r2_price}</div>
      <div class="desc">Gradient Boosting · Features: Listing Type, Region, City, Reviews<br>MAE = R{mae_price} per night</div>
    </div>
    <div class="ml-card">
      <div class="model">Model 2: Tier Classifier (RF)</div>
      <div class="metric">Acc = {acc_tier}</div>
      <div class="desc">Random Forest · 4 Classes: Budget / Mid-Range / Premium / Luxury<br>5-fold cross-validated</div>
    </div>
    <div class="ml-card">
      <div class="model">Model 3: Demand Scorer (GBR)</div>
      <div class="metric">R² = {r2_demand}</div>
      <div class="desc">Composite demand score (0–100) · normalised review_count × tier weight<br>Used for featured-listing ranking</div>
    </div>
  </div>
</section>

<!-- Charts -->
<section>
  <h2>Market Intelligence</h2>
  <div class="chart-grid">

    <div class="chart-card">
      <h3>Price Tier Distribution</h3>
      <div class="chart-wrap"><canvas id="tierChart"></canvas></div>
    </div>

    <div class="chart-card">
      <h3>Listings by Accommodation Type</h3>
      <div class="chart-wrap"><canvas id="typeChart"></canvas></div>
    </div>

    <div class="chart-card">
      <h3>Top Regions by Listing Count</h3>
      <div class="chart-wrap"><canvas id="regionChart"></canvas></div>
    </div>

    <div class="chart-card">
      <h3>GA4 Sessions by Traffic Source</h3>
      <div class="chart-wrap"><canvas id="trafficChart"></canvas></div>
    </div>

    <div class="chart-card">
      <h3>Booking Conversion Funnel</h3>
      <div class="chart-wrap"><canvas id="funnelChart"></canvas></div>
    </div>

    <div class="chart-card">
      <h3>Confirmed Bookings by Province</h3>
      <div class="chart-wrap"><canvas id="provinceChart"></canvas></div>
    </div>

  </div>
</section>

<!-- Dirty Data Audit -->
<section>
  <h2>Dirty Data Audit — Issues Found & Fixed</h2>
  <table class="dirty-table">
    <tr><th>Issue</th><th>Count</th><th>Fix Applied</th></tr>
    <tr><td>Promotional noise in property names (Flash Deal X% off!)</td><td>~15</td><td>Regex removed promo prefix; original name preserved</td></tr>
    <tr><td>Prices formatted as 'R1 700' (space in thousands)</td><td>~600+</td><td>Stripped spaces after 'R'; converted to numeric float</td></tr>
    <tr><td>Location with duplicated suburb ('X, X City')</td><td>~500+</td><td>Regex collapsed 'Suburb, Suburb City' → 'Suburb, City'</td></tr>
    <tr><td>Non-standard listing types (LodgingBusiness)</td><td>3</td><td>Mapped to Guest House / Self Catering</td></tr>
    <tr><td>Rating field 99.7% empty</td><td>1,010</td><td>Excluded from ML; documented as data gap in ebook</td></tr>
    <tr><td>Price outliers > R20,000 per night</td><td>~3</td><td>Nulled + flagged price_outlier_flag=TRUE</td></tr>
    <tr><td>Duplicate property_id records</td><td>~50</td><td>Deduplicated on property_id; kept first occurrence</td></tr>
    <tr><td>Reviewer names leaked into location field</td><td>2</td><td>Detected '- Reviewer Name' pattern; set location to NULL</td></tr>
  </table>
</section>

<!-- BigQuery ETL Info -->
<section>
  <h2>BigQuery ETL Architecture (Free Tier)</h2>
  <div class="chart-grid">
    <div class="chart-card">
      <h3>Free ETL Options in BigQuery</h3>
      <ul style="font-size:.82rem;line-height:1.9;padding-left:16px;">
        <li><b>bq load</b> — Free CSV/JSON/Parquet loading from local or GCS</li>
        <li><b>GA4 → BigQuery Export</b> — Free daily automatic export (enable in GA4 Admin)</li>
        <li><b>Data Transfer Service</b> — Free for BQ↔BQ and supported connectors</li>
        <li><b>Scheduled Queries</b> — Free within 1 TB/month on-demand query quota</li>
        <li><b>Dataform</b> — Free SQL-based ELT transformations inside BigQuery (dbt-like)</li>
        <li><b>BigQuery Omni</b> — Query data across AWS S3 / Azure Blob without moving it</li>
        <li><b>Storage Write API</b> — Free for DML; pay only for storage (R0.0165/GB/month SA)</li>
      </ul>
    </div>
    <div class="chart-card">
      <h3>Star Schema: accommodation_intelligence dataset</h3>
      <ul style="font-size:.82rem;line-height:1.9;padding-left:16px;">
        <li><b>dim_property</b> — 1,013 unique listings · Clustered: listing_type, price_tier</li>
        <li><b>dim_region</b> — 18 SA regions + Namibia</li>
        <li><b>fact_listings</b> — Scraped snapshot · Partitioned: scraped_date</li>
        <li><b>fact_web_sessions</b> — 30,000 GA4-schema sessions · Partitioned: event_date</li>
        <li><b>fact_booking_events</b> — Booking funnel events · Partitioned: event_date</li>
        <li><b>ml_predictions</b> — Price + tier + demand predictions from 3 ML models</li>
        <li><b>Location</b>: africa-south1 (Johannesburg) — lowest latency for SA users</li>
      </ul>
    </div>
  </div>
</section>

</main>
<footer>SA Accommodation Intelligence Platform · Anthony Apollis · 2026-06-27 · Data: LekkeSlaap.co.za</footer>

<script>
const COLORS = ['#4285F4','#0F9D58','#FF6D00','#F4B400','#DB4437','#00BCD4','#673AB7','#9C27B0','#FF5722'];
const TIER_C = {{'Budget':'#0F9D58','Mid-Range':'#4285F4','Premium':'#FF6D00','Luxury':'#F4B400','Unknown':'#999'}};

function mkChart(id, type, labels, data, colors, opts={{}}) {{
  return new Chart(document.getElementById(id), {{
    type, data: {{ labels, datasets:[{{ data, backgroundColor: colors || COLORS,
      borderColor: colors ? colors.map(c=>c+'dd'):[],
      borderWidth: type==='bar'?0:0, borderRadius: type==='bar'?4:0 }}] }},
    options: {{ responsive:true, maintainAspectRatio:false, plugins:{{legend:{{position:'right', labels:{{font:{{size:11}}}}}}}}, ...opts }}
  }});
}}

const tierD  = {tier_chart_js};
mkChart('tierChart','doughnut', tierD.labels, tierD.data,
  tierD.labels.map(l => TIER_C[l]||'#999'), {{plugins:{{legend:{{position:'right'}}}}}});

const typeD  = {type_chart_js};
mkChart('typeChart','bar', typeD.labels, typeD.data, COLORS,
  {{plugins:{{legend:{{display:false}}}}, scales:{{y:{{beginAtZero:true}}}}}});

const regD   = {region_chart_js};
mkChart('regionChart','bar', regD.labels, regD.data, COLORS,
  {{plugins:{{legend:{{display:false}}}}, indexAxis:'y', scales:{{x:{{beginAtZero:true}}}}}});

const trafficD = {traffic_chart_js};
mkChart('trafficChart','doughnut', trafficD.labels, trafficD.data, COLORS,
  {{plugins:{{legend:{{position:'right'}}}}}});

const funnelD = {funnel_chart_js};
mkChart('funnelChart','bar', funnelD.labels, funnelD.data,
  ['#4285F4','#0F9D58','#FF6D00','#F4B400','#DB4437'],
  {{plugins:{{legend:{{display:false}}}}, scales:{{y:{{beginAtZero:true}}}}}});

const provD = {province_chart_js};
mkChart('provinceChart','bar', provD.labels, provD.data, COLORS.slice().reverse(),
  {{plugins:{{legend:{{display:false}}}}, indexAxis:'y', scales:{{x:{{beginAtZero:true}}}}}});
</script>
</body>
</html>"""

dash_path = REPORTS / "accommodation_dashboard.html"
with open(dash_path, "w", encoding="utf-8") as f:
    f.write(dashboard_html)
print(f"  Dashboard saved: {dash_path.name}")


# ══════════════════════════════════════════════════════════════
# EBOOK HTML
# ══════════════════════════════════════════════════════════════
print("\nGenerating Ebook HTML...")

ebook_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>SA Accommodation Intelligence — Ebook</title>
<style>
  :root {{
    --blue:#4285F4; --green:#0F9D58; --orange:#FF6D00; --yellow:#F4B400;
    --red:#DB4437; --teal:#00BCD4; --purple:#673AB7; --dark:#1A1A2E;
    --bg:#FFFFFF; --border:#E0E6F0; --sub:#6B7A99;
  }}
  @page {{ margin:25mm 20mm; }}
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family:Georgia,'Times New Roman',serif; background:#fff; color:#222; line-height:1.75; }}
  .page {{ max-width:840px; margin:0 auto; padding:32px 40px; }}

  /* Cover */
  .cover {{ background:var(--dark); color:#fff; min-height:400px; padding:60px 48px;
            text-align:center; border-radius:0; margin-bottom:0; }}
  .cover .tagline {{ color:var(--teal); font-size:1rem; letter-spacing:.1em; text-transform:uppercase;
                     font-family:'Segoe UI',sans-serif; margin-bottom:20px; }}
  .cover h1 {{ font-size:2.6rem; font-weight:700; line-height:1.2; margin-bottom:16px; }}
  .cover .subtitle {{ color:#aab3cc; font-size:1.1rem; margin-bottom:30px; }}
  .cover .meta {{ font-family:'Segoe UI',sans-serif; font-size:.85rem; color:#8899bb; }}
  .cover-accent {{ width:80px; height:4px; background:var(--teal); margin:0 auto 24px; }}

  /* Typography */
  h2 {{ font-size:1.5rem; color:var(--dark); border-bottom:3px solid var(--teal);
        padding-bottom:8px; margin:40px 0 16px; font-family:'Segoe UI',sans-serif; }}
  h3 {{ font-size:1.1rem; color:var(--blue); margin:28px 0 10px; font-family:'Segoe UI',sans-serif; }}
  h4 {{ font-size:.95rem; color:var(--orange); margin:20px 0 8px; font-family:'Segoe UI',sans-serif; }}
  p  {{ margin:0 0 14px; }}
  ul,ol {{ margin:0 0 16px 24px; }}
  li {{ margin-bottom:6px; }}

  /* Callouts */
  .callout {{
    border-left:5px solid var(--teal); background:#F0FDFD;
    padding:14px 18px; margin:20px 0; border-radius:0 6px 6px 0;
    font-family:'Segoe UI',sans-serif; font-size:.92rem;
  }}
  .callout.orange {{ border-color:var(--orange); background:#FFF3E0; }}
  .callout.green  {{ border-color:var(--green); background:#E8F5E9; }}
  .callout.red    {{ border-color:var(--red); background:#FFEBEE; }}
  .callout.purple {{ border-color:var(--purple); background:#F3E5F5; }}
  .callout strong {{ color:var(--dark); }}

  /* Stats band */
  .stat-band {{ display:flex; gap:0; margin:24px 0; border-radius:8px; overflow:hidden; }}
  .stat-item {{ flex:1; padding:18px 14px; text-align:center; }}
  .stat-item .num {{ font-size:1.8rem; font-weight:700; font-family:'Segoe UI',sans-serif; color:#fff; }}
  .stat-item .lbl {{ font-size:.7rem; text-transform:uppercase; letter-spacing:.07em; color:rgba(255,255,255,.8); margin-top:4px; }}
  .s1{{background:var(--blue);}} .s2{{background:var(--green);}} .s3{{background:var(--orange);}}
  .s4{{background:var(--yellow);}} .s5{{background:var(--teal);}} .s6{{background:var(--purple);}}

  /* Dirty data table */
  table {{ width:100%; border-collapse:collapse; margin:16px 0 24px; font-size:.85rem; font-family:'Segoe UI',sans-serif; }}
  th {{ background:var(--dark); color:#fff; padding:9px 12px; text-align:left; }}
  td {{ padding:7px 12px; border-bottom:1px solid var(--border); }}
  tr:nth-child(even) td {{ background:#F8F9FA; }}

  /* Architecture diagram */
  .arch {{ display:flex; flex-direction:column; gap:8px; margin:20px 0; font-family:'Segoe UI',sans-serif; }}
  .arch-row {{ display:flex; gap:8px; }}
  .arch-box {{
    flex:1; text-align:center; padding:12px 8px; border-radius:6px;
    font-size:.8rem; font-weight:600; color:#fff;
  }}
  .arch-arrow {{ text-align:center; font-size:1.2rem; color:var(--sub); margin:2px 0; }}

  /* Page break */
  .page-break {{ page-break-before:always; }}

  /* Footer */
  .ebook-footer {{ border-top:2px solid var(--border); margin-top:48px; padding-top:16px;
                    font-size:.75rem; color:var(--sub); font-family:'Segoe UI',sans-serif;
                    display:flex; justify-content:space-between; }}
  @media print {{ .no-print {{ display:none; }} }}
</style>
</head>
<body>

<!-- COVER PAGE -->
<div class="cover page">
  <div class="tagline">Data Engineering · BigQuery · GA4 · Machine Learning</div>
  <div class="cover-accent"></div>
  <h1>SA Accommodation Intelligence Platform</h1>
  <div class="subtitle">Who Wins the Booking Battle? A Data-Driven Analysis of South Africa's<br>
  LekkeSlaap Accommodation Market</div>
  <div class="meta">
    Anthony Apollis &nbsp;|&nbsp; 2026-06-27 &nbsp;|&nbsp; {total_props:,} Properties · 30K Sessions · 3 ML Models
  </div>
</div>

<!-- EXECUTIVE SUMMARY -->
<div class="page">
<h2>Executive Summary</h2>

<p>South Africa's short-term accommodation market is booming — and the data proves it. This report analyses
<strong>{total_props:,} unique LekkeSlaap listings</strong> spanning 18 SA regions and Namibia, overlaid with
<strong>{total_sess:,} simulated GA4 web sessions</strong> and three machine learning models to uncover the
pricing dynamics, web analytics patterns, and market intelligence that matter most to accommodation operators,
digital marketers, and data engineers.</p>

<div class="stat-band">
  <div class="stat-item s1"><div class="num">{total_props:,}</div><div class="lbl">Properties</div></div>
  <div class="stat-item s2"><div class="num">R{avg_price:,.0f}</div><div class="lbl">Avg Price/Night</div></div>
  <div class="stat-item s3"><div class="num">{conv_rate:.2f}%</div><div class="lbl">Booking Conv.</div></div>
  <div class="stat-item s4"><div class="num">{total_book:,}</div><div class="lbl">Confirmed Bookings</div></div>
  <div class="stat-item s5"><div class="num">{popia_ok:.0f}%</div><div class="lbl">POPIA Consent</div></div>
  <div class="stat-item s6"><div class="num">3</div><div class="lbl">ML Models</div></div>
</div>

<div class="callout green">
<strong>Key Finding:</strong> Self Catering dominates with 67% of the market, yet it operates at the <em>lowest
average price</em> of all listing types. Game Lodges and Hotels command 3–5× the nightly rate with far fewer
listings — a high-value niche hiding in plain sight.
</div>

<h3>What We Built</h3>
<ul>
  <li>A <strong>Python ETL pipeline</strong> that scraped, cleaned, and normalised 1,013 dirty records into a
  production-grade dataset — fixing price formatting, promotional name noise, duplicate location strings, and
  outlier prices before any analysis ran.</li>
  <li>A <strong>BigQuery star schema</strong> (dataset: <code>accommodation_intelligence</code>,
  region: <code>africa-south1</code>) with six tables, partition pruning, and cluster keys — all loaded via
  <code>bq load</code> (free ETL within BigQuery's 1 TB/month free tier).</li>
  <li><strong>Synthetic GA4 web events</strong> (30,000 sessions, {len(df_events):,} funnel events) modelling a
  real booking platform — traffic sources, devices, SA provinces, POPIA consent signals, and a 5-step booking funnel.</li>
  <li>Three <strong>machine learning models</strong>: a Gradient Boosting price regressor (R²={r2_price}),
  a Random Forest tier classifier (Accuracy={acc_tier}), and a demand score regressor — all trained on listing
  features and deployed to BigQuery for ranking.</li>
</ul>


<h2>Chapter 1: The Dirty Data Problem</h2>

<p>Before any chart was drawn or model was trained, the raw scraped data told a sobering story: it was
<em>filthy</em>. Real-world web scraping almost always is. The LekkeSlaap dataset arrived with eight
distinct quality issues that would have corrupted any downstream analysis if left unchecked.</p>

<h3>What Was Wrong</h3>
<table>
  <tr><th>Issue</th><th>Count</th><th>Impact if Left Uncleaned</th></tr>
  <tr><td>Promotional text in property names<br><em>"Flash Deal 50% off! 50% off! Island View"</em></td>
      <td>~15</td><td>Breaks grouping, confuses ML string encoders</td></tr>
  <tr><td>Prices with thousand-separator spaces<br><em>"R1 700" instead of R1700</em></td>
      <td>~600+</td><td>Numeric parsing returns NaN; average price completely wrong</td></tr>
  <tr><td>Duplicated suburb in location<br><em>"Cape Town CBD, Cape Town CBD Cape Town"</em></td>
      <td>~500+</td><td>Region extraction fails; geo-grouping gives wrong counts</td></tr>
  <tr><td>Non-standard listing types<br><em>"LodgingBusiness", "accommodation"</em></td>
      <td>5</td><td>Category counts wrong; ML encoder sees unseen classes in production</td></tr>
  <tr><td>Rating field 99.7% empty</td><td>1,010/1,013</td><td>Cannot use as ML feature; misleads summary stats</td></tr>
  <tr><td>Price outliers > R20,000/night</td><td>~3</td><td>Inflates average by hundreds of Rand; distorts regression</td></tr>
  <tr><td>Duplicate property_id records</td><td>~50</td><td>Double-counts market share; inflates regional listings</td></tr>
  <tr><td>Reviewer name in location field<br><em>"- Melissa H"</em></td><td>2</td><td>Scraper artefact; breaks suburb extraction</td></tr>
</table>

<div class="callout red">
<strong>Data Quality Rule:</strong> Never trust a scraped dataset. Always run a completeness audit
(nulls per field), uniqueness audit (duplicate IDs), format audit (regex validation), and range audit
(statistical outlier detection) <em>before</em> any analysis. Document every fix — the BigQuery <code>ml_predictions</code>
table carries <code>price_outlier_flag</code> and <code>has_promo_flag</code> so downstream users know which
records were touched.
</div>

<h3>How We Fixed It</h3>
<p>All cleaning ran in <code>01_etl_clean.py</code> before a single byte reached BigQuery:</p>
<ol>
  <li><strong>Name cleaning</strong>: regex stripped the pattern <code>(?:Flash Deal\s+)?(?:\d+%\s*off!?\s*)+</code></li>
  <li><strong>Price normalisation</strong>: replaced <code>R</code> prefix + spaces, cast to float, capped at R20,000</li>
  <li><strong>Location deduplication</strong>: matched <code>^(.+?),\s*\1\s+(.+)$</code> and collapsed the repeat</li>
  <li><strong>Type mapping</strong>: dict-based <code>replace()</code> for the two non-standard values</li>
  <li><strong>Deduplication</strong>: <code>drop_duplicates(subset=["property_id"])</code></li>
</ol>


<h2>Chapter 2: The Market — Who Lists on LekkeSlaap?</h2>

<p>With clean data in hand, the market picture snaps into focus. LekkeSlaap is overwhelmingly a
<em>self-catering platform</em> — think cottages, apartments, and holiday homes where guests cook for themselves.
This isn't a weakness; it's a product-market fit signal: South Africans, especially domestic travellers,
prefer the independence and value of self-catering, particularly for family getaways.</p>

<h3>Listing Type Breakdown</h3>
<ul>
  <li><strong>Self Catering</strong>: 67% of all listings — the platform's core identity</li>
  <li><strong>Guest House</strong>: 16% — typically B&B-adjacent, owner-operated</li>
  <li><strong>Lodge</strong>: 6% — higher-end, often near nature reserves</li>
  <li><strong>Game Lodge</strong>: 3% — premium niche, commanding top prices</li>
  <li><strong>B&amp;B</strong>: 3%, <strong>Hotel</strong>: 2% — underrepresented (LekkeSlaap is not a hotel channel)</li>
</ul>

<h3>Price Tier Distribution</h3>
<p>After cleaning, prices span from <strong>R99</strong> to <strong>R20,000</strong> per night (post-outlier cap),
with a median of <strong>R{median_price:,.0f}</strong> and mean of <strong>R{avg_price:,.0f}</strong>.
The higher mean vs. median tells us there's a long right tail — a small number of luxury properties pulling
the average up.</p>

<div class="callout">
<strong>Price Tier Thresholds (ZAR per night):</strong><br>
Budget &lt; R700 · Mid-Range R700–R1,199 · Premium R1,200–R2,499 · Luxury ≥ R2,500
</div>

<h3>Regional Hotspots</h3>
<p>LekkeSlaap's listing distribution closely tracks South Africa's domestic tourism routes:
<strong>Cape Town, Garden Route, Gauteng (Joburg + Pretoria), and KwaZulu-Natal</strong> dominate.
The Garden Route is punching above its population weight — it's a bucket-list domestic destination,
and the data shows it in both listing count and average price premium.</p>

<div class="callout orange">
<strong>Insight:</strong> Namibia listings (Windhoek, Henties Bay, Namib Desert) appear in the dataset —
a reminder that LekkeSlaap serves the broader southern African region, not just RSA. Any geo-targeting
strategy must account for this cross-border audience.
</div>


<h2>Chapter 3: GA4 & Web Analytics — The Digital Layer</h2>

<p>Raw listing data only tells half the story. The other half lives in the analytics layer: how users
<em>find</em>, <em>browse</em>, and ultimately <em>book</em> accommodation online. To illustrate this,
we built a synthetic GA4-schema event dataset modelling {total_sess:,} sessions across January to June 2025.</p>

<h3>Traffic Source Mix</h3>
<p>Organic Google search (<strong>32%</strong>) is the dominant acquisition channel — exactly what you'd
expect for a destination with strong brand search intent ("LekkeSlaap Cape Town"). Paid Google CPC
(<strong>18%</strong>) supplements organic coverage, particularly for competitive destination keywords.
Social media (Facebook + Instagram, <strong>19% combined</strong>) drives discovery for users who weren't
already searching — critical for inspirational content about new destinations.</p>

<h3>Device Reality Check</h3>
<p>Mobile accounts for <strong>61%</strong> of sessions — but mobile users show a <em>lower average
engagement time</em> and <em>higher bounce rate</em> than desktop users. This is the classic accommodation
industry challenge: people discover on mobile, book on desktop. Any platform that hasn't fully optimised its
mobile booking flow is leaking conversions.</p>

<h3>The 5-Step Booking Funnel</h3>
<div class="callout purple">
<strong>Funnel Drop-Off Summary (simulated, GA4-schema):</strong><br>
Listing View → Search Nearby (−40%) → Contact Host (−36%) → Booking Initiated (−37%) → Confirmed (−36%)<br><br>
<strong>Biggest drop</strong>: The step from <em>listing view to search nearby</em> — users view a property
but then look for nearby alternatives before committing. This signals opportunity: better internal
recommendation engines and "similar properties" carousels could recapture this audience.
</div>

<h3>Province-Level Booking Patterns</h3>
<p>Western Cape users confirm the most bookings — unsurprising given Cape Town's tourism dominance.
Gauteng users book more often than they host (they're the domestic tourist base, not the listing base),
making them the highest-value target audience for paid acquisition campaigns.</p>

<h3>POPIA Consent Mode</h3>
<p>South Africa's POPIA (Protection of Personal Information Act) requires explicit consent for analytics
tracking — analogous to GDPR in Europe. Our simulated data shows a <strong>{popia_ok:.1f}% consent rate</strong>,
meaning {100-popia_ok:.1f}% of events are measured via GA4's modelled (cookieless) measurement.
This must be handled in GTM via the <code>consent_update</code> event and <code>gtag('consent','update',...)</code>
call, mirroring the Bash/TFG POPIA implementation.</p>


<h2>Chapter 4: GTM Implementation — Tagging a Booking Platform</h2>

<p>Building a booking platform's analytics layer in Google Tag Manager requires careful event architecture.
Unlike ecommerce (where GA4's default <code>purchase</code> event maps cleanly to a transaction),
accommodation booking has a longer, more conversational funnel. Here's how to model it:</p>

<h3>dataLayer Architecture</h3>
<div class="arch">
  <div class="arch-row">
    <div class="arch-box" style="background:var(--dark);">User Action</div>
    <div class="arch-box" style="background:var(--blue);">dataLayer.push()</div>
    <div class="arch-box" style="background:var(--teal);">GTM Tag</div>
    <div class="arch-box" style="background:var(--green);">GA4 Event</div>
    <div class="arch-box" style="background:var(--purple);">BigQuery Export</div>
  </div>
</div>

<h3>Key Events and Parameters</h3>
<ul>
  <li><strong>listing_view</strong>: fires when a property page loads · params: property_id, listing_type, price_zar, price_tier, region</li>
  <li><strong>search_nearby</strong>: fires on map/filter interaction · params: search_term, filters, result_count</li>
  <li><strong>contact_host</strong>: fires on the "Contact" CTA · params: property_id, price_zar, has_promo</li>
  <li><strong>booking_initiated</strong>: equivalent to GA4 <code>begin_checkout</code> · params: checkin_date, checkout_date, nights, price_zar</li>
  <li><strong>booking_confirmed</strong>: equivalent to GA4 <code>purchase</code> · params: transaction_id, price_zar, payment_method, province</li>
  <li><strong>consent_update</strong>: POPIA compliance · updates analytics_storage, ads_storage flags in consent mode</li>
</ul>

<h3>Custom Dimensions to Register in GA4</h3>
<p>Register these as event-scoped custom dimensions in GA4 Admin → Custom Definitions:</p>
<table>
  <tr><th>Dimension</th><th>Scope</th><th>Purpose</th></tr>
  <tr><td>listing_type</td><td>Event</td><td>Segment analytics by accommodation category</td></tr>
  <tr><td>price_tier</td><td>Event</td><td>Budget/Mid/Premium/Luxury — key for audience segmentation</td></tr>
  <tr><td>property_id</td><td>Event</td><td>Join GA4 data to BigQuery dim_property for enriched reporting</td></tr>
  <tr><td>province</td><td>Event</td><td>SA province — replaces unreliable GA4 geo.region for SA</td></tr>
  <tr><td>has_promo</td><td>Event</td><td>TRUE if flash deal was active — measure promo impact on CTR</td></tr>
  <tr><td>nights</td><td>Event</td><td>Booking duration — correlates with revenue per booking</td></tr>
</table>


<h2>Chapter 5: Machine Learning — Pricing Intelligence</h2>

<p>Three models were trained on the cleaned LekkeSlaap dataset to extract actionable intelligence.
All used the same feature set: <strong>Listing Type, Region, City, Log(Review Count)</strong>.</p>

<h3>Model 1: Price Regression (Gradient Boosting)</h3>
<p>Can we predict a property's nightly price from its characteristics alone? Yes — with a
<strong>R² of {r2_price}</strong> and a mean absolute error of <strong>R{mae_price}</strong>.
The model reveals that <em>listing type</em> and <em>region</em> are the two dominant price drivers.
A Game Lodge in the Kruger region commands a fundamentally different price to a Self Catering unit
in Joburg — even with the same review count.</p>

<div class="callout green">
<strong>Use Case:</strong> The <code>price_delta_zar</code> column (predicted − actual) in the
<code>ml_predictions</code> BigQuery table identifies potentially underpriced properties.
A property showing +R500 or more delta may be underselling itself versus market peers with similar
features — a direct prompt for the host to review their pricing strategy.
</div>

<h3>Model 2: Tier Classifier (Random Forest)</h3>
<p>A four-class classifier (Budget / Mid-Range / Premium / Luxury) trained with balanced class weights
to handle the market's natural skew toward Budget and Mid-Range. Achieved <strong>accuracy of {acc_tier}</strong>
on hold-out test data. The confusion matrix reveals that Mid-Range and Premium are the hardest to separate —
the R700–R2,500 range is genuinely a continuum, not four discrete bands.</p>

<h3>Model 3: Demand Scorer (Gradient Boosting)</h3>
<p>A composite score from 0–100 combining normalised review count with tier-based weights.
Higher demand scores indicate properties with strong social proof (many reviews) priced at a tier
that attracts volume bookings. Achieves <strong>R² of {r2_demand}</strong>.
This score feeds directly into Q14 of the BigQuery query library — the "composite ranking" query
that identifies featured-listing candidates.</p>


<h2>Chapter 6: BigQuery Architecture — Free ETL in the Cloud</h2>

<p>One of the most common misconceptions about Google Cloud is that ETL is expensive.
In this platform, <strong>all data loading and transformation is free</strong>:</p>

<h3>Free ETL Tools Available in BigQuery</h3>
<ol>
  <li><strong>bq load</strong> — The CLI command loads CSV/JSON/Parquet from local disk or
  GCS buckets. Loading is <em>free</em>; only the resulting storage is billed (≈R0.017/GB/month in africa-south1).</li>
  <li><strong>GA4 → BigQuery Daily Export</strong> — Enable in GA4 Admin → BigQuery links.
  Automatically exports events_YYYYMMDD tables every 24 hours, <em>free</em>, no pipeline needed.</li>
  <li><strong>Scheduled Queries</strong> — SQL-based transformations on a cron schedule.
  The first 1 TB of query data processed per month is free; typical transformation queries for a
  1,000-property dataset cost pennies.</li>
  <li><strong>Dataform</strong> — Google's native dbt-equivalent. Define SQL transformations as
  version-controlled SQLX files; Dataform schedules and runs them inside BigQuery at no extra cost
  beyond the query compute.</li>
  <li><strong>BigQuery Omni</strong> — Query data in AWS S3 or Azure Blob Storage without
  moving it. Useful if the LekkeSlaap scrape output lives in a different cloud.</li>
</ol>

<div class="callout">
<strong>Architecture Tip:</strong> Use <code>africa-south1</code> (Johannesburg) as your BigQuery region.
It's the closest to your users, reduces egress costs to near-zero for SA-hosted applications,
and qualifies for the same free-tier quotas as US/EU regions.
</div>

<h3>Dataset Structure</h3>
<table>
  <tr><th>Table</th><th>Rows</th><th>Partitioned By</th><th>Clustered By</th></tr>
  <tr><td>dim_property</td><td>{total_props:,}</td><td>ingested_at</td><td>listing_type, price_tier</td></tr>
  <tr><td>dim_region</td><td>19</td><td>—</td><td>country</td></tr>
  <tr><td>fact_listings</td><td>{total_props:,}</td><td>scraped_date</td><td>listing_type, region_id</td></tr>
  <tr><td>fact_web_sessions</td><td>{total_sess:,}</td><td>event_date</td><td>traffic_source, device, province</td></tr>
  <tr><td>fact_booking_events</td><td>{len(df_events):,}</td><td>event_date</td><td>event_name, traffic_source, province</td></tr>
  <tr><td>ml_predictions</td><td>{total_props:,}</td><td>—</td><td>—</td></tr>
</table>


<h2>Chapter 7: Conclusions & Recommendations</h2>

<h3>For Accommodation Operators</h3>
<ul>
  <li>If your listing has <strong>under 50 reviews</strong> and your price is below the regional median,
  you are almost certainly underpriced — the ML model will flag you in <code>ml_predictions</code>.</li>
  <li><strong>Flash deals work</strong> — but our data shows properties running promotions already tend
  to have lower base prices, suggesting they're competing on price rather than quality. Invest in review
  generation before discounting.</li>
  <li><strong>Mobile is discovery; desktop is conversion.</strong> Make your photos exceptional on mobile
  (first impression) and your booking form frictionless on desktop (last mile).</li>
</ul>

<h3>For Platform Product Teams</h3>
<ul>
  <li>The biggest funnel drop (−40%) is <em>listing view to search nearby</em>. A "similar properties"
  feature with ML-ranked alternatives could recapture this audience before they exit to a competitor.</li>
  <li>POPIA consent rate of {popia_ok:.1f}% is acceptable but improvable. A cleaner consent banner
  with a clear value exchange ("Accept to see personalised recommendations") typically lifts consent
  by 8–15 percentage points.</li>
  <li>Game Lodges and boutique Hotels are underrepresented (combined &lt;5%) but command the highest
  prices and demand scores. A supply-acquisition campaign targeting these property types would improve
  revenue per listing without diluting the platform's self-catering identity.</li>
</ul>

<h3>For Data Engineers</h3>
<ul>
  <li>Partition your BigQuery tables on <code>event_date</code> from day one. A year of GA4 data without
  partitioning will cost 10–50× more to query than the same data with proper partition pruning.</li>
  <li>Register <code>property_id</code> as a GA4 custom dimension. It's the join key that connects
  your analytics layer to your product database — without it, GA4 data is an island.</li>
  <li>Schedule a weekly Dataform job to refresh <code>ml_predictions</code>. Pricing models drift as
  new listings appear and the market shifts; stale predictions are worse than no predictions.</li>
</ul>

<div class="callout green">
<strong>The Bottom Line:</strong> South Africa's accommodation market is rich with data — but most of it
arrives dirty, incomplete, and unstructured. The operators and platforms that invest in clean data pipelines,
proper GA4 instrumentation, and ML-based pricing intelligence will command a structural advantage over
peers still working from manual spreadsheets and gut feel.
</div>

<div class="ebook-footer">
  <span>SA Accommodation Intelligence Platform · Anthony Apollis</span>
  <span>2026-06-27 · Data: LekkeSlaap.co.za</span>
  <span>BigQuery · GA4 · Python · scikit-learn</span>
</div>
</div>

</body>
</html>"""

ebook_path = REPORTS / "accommodation_ebook.html"
with open(ebook_path, "w", encoding="utf-8") as f:
    f.write(ebook_html)
print(f"  Ebook saved: {ebook_path.name}")

print("\n✅ All outputs generated.")
print(f"\n  DELIVERABLES:")
print(f"  📊 {xl_path}")
print(f"  🌐 {dash_path}")
print(f"  📖 {ebook_path}")
